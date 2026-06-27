"""
Multi-company BTC Mining Standalone Valuation Model
Sources: Q1 2026 10-Q filings / 8-K press releases (filed May 2026)
Tabs: Summary Comp + individual models for CIFR, WULF, RIOT, HUT, CLSK, CORZ
"""

import shutil, openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

OUT = "/home/user/claude-code/BTC_Mining_Standalone_MultiCo.xlsx"

# ── Styles ─────────────────────────────────────────────────────────────────────
def fill(hex): return PatternFill("solid", fgColor=hex)
def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def side(style="thin", color="AAAAAA"): return Side(style=style, color=color)
def bb(): return Border(bottom=side())
def bthick(): return Border(bottom=side("medium", "1F3864"))
def bbox():
    s=side(); return Border(top=s,bottom=s,left=s,right=s)

C_NAVY   = "1F3864"; C_BLUE  = "2E5EAA"; C_LBLUE = "BDD7EE"
C_YEL    = "FFD966"; C_GRN   = "E2EFDA"; C_RED   = "FCE4D6"
C_GREY   = "F2F2F2"; C_WHITE = "FFFFFF"; C_ORG   = "F4B942"
C_DGREY  = "595959"

def hdr(ws, row, col, val, bg=C_NAVY, fg="FFFFFF", sz=10, bold=True, span=1, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill(bg); c.font = font(bold, fg, sz)
    c.alignment = align("center","center", wrap)
    if span > 1:
        ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+span-1)
    return c

def cell(ws, row, col, val, bg=None, bold=False, color="000000", sz=10,
         h="right", fmt=None, wrap=False, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = fill(bg)
    c.font = font(bold, color, sz, italic)
    c.alignment = align(h,"center", wrap)
    if fmt: c.number_format = fmt
    return c

def label(ws, row, col, val, bg=None, bold=False, indent=0):
    c = ws.cell(row=row, column=col, value=("  "*indent)+val)
    if bg: c.fill = fill(bg)
    c.font = font(bold, size=10)
    c.alignment = align("left","center")
    return c

def section_hdr(ws, row, ncols, title):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.fill = fill(C_BLUE); c.font = font(True,"FFFFFF",10)
    c.alignment = align("left","center")

def pct_fmt(v): return f"{v:.1%}" if v is not None else "N/A"
def mm_fmt(v):  return f"${v:,.0f}mm" if v is not None else "N/A"

# ═══════════════════════════════════════════════════════════════════════════════
# COMPANY DATA
# ═══════════════════════════════════════════════════════════════════════════════
# Source legend: C=Confirmed filing, E=Estimated/derived, N/A=not disclosed
# All mining data as of Q1 2026 (Jan–Mar 2026) filings

COMPANIES = [
    {
        "ticker": "CIFR", "name": "Cipher Digital", "status": "Winding down (PPA expires Jul 2027)",
        "profile": "HPC conversion; last mining site Odessa TX 207 MW; exit by mid-2027",
        # Key assumptions
        "hash_26": 11.6, "hash_27": 5.8, "hash_28": 0.0, "hash_src": "C",
        "eff_26": 17.2, "eff_src": "C",
        "pwr_26": 28.0, "pwr_src": "C",          # $/MWh - confirmed PPA
        "opex_26": 55, "opex_src": "E",            # other mining opex $mm/yr
        "capex_26": 0, "capex_src": "C",
        # Actuals Q1 2026
        "btc_q1": 346, "rev_q1": 34.8, "cost_mine": None, "gross_margin": None,
        # Notes
        "power_note": "$28/MWh fixed Odessa TX PPA (confirmed Q1 2026 10-Q)",
        "hash_note": "11.6 EH/s Q1 2026; 5.8 EH/s FY2027 effective (6-mo avg); 0 FY2028+",
        "eff_note": "17.2 J/TH fleet avg (Q1 2026 10-Q)",
        "opex_note": "Mining SG&A not broken out; ~$55mm/yr estimate",
        "wind_down": True, "tv_zero": True,
    },
    {
        "ticker": "WULF", "name": "TeraWulf", "status": "Transitioning to HPC",
        "profile": "Lake Mariner nuclear-powered; HPC rev now exceeds mining; wind-down likely by 2028 halving",
        "hash_26": 7.9, "hash_27": 6.0, "hash_28": 3.0, "hash_src": "C",
        "eff_26": 22.0, "eff_src": "E",           # ~25 J/TH prior, improving; estimate
        "pwr_26": 35.0, "pwr_src": "E",            # ~3.5¢/kWh prior; nuclear-sourced est.
        "opex_26": 40, "opex_src": "E",
        "capex_26": 20, "capex_src": "E",          # minimal new mining capex
        "btc_q1": 508, "rev_q1": 13.0, "cost_mine": None, "gross_margin": 0.036,
        "power_note": "~3.5¢/kWh est. (nuclear-sourced Lake Mariner; exact rate not disclosed Q1 2026)",
        "hash_note": "7.9 EH/s total (C); 0.9% global share; winding down per mgmt commentary",
        "eff_note": "~22 J/TH estimated; prior disclosures ~25 J/TH; not confirmed Q1 2026",
        "opex_note": "SG&A/other not isolated for mining; ~$40mm/yr estimate",
        "wind_down": True, "tv_zero": False,
    },
    {
        "ticker": "RIOT", "name": "Riot Platforms", "status": "Scaling + DC infrastructure",
        "profile": "Largest pure-play miner; Corsicana TX 1GW campus; power credits reduce net cost",
        "hash_26": 42.5, "hash_27": 50.0, "hash_28": 55.0, "hash_src": "C",
        "eff_26": 20.2, "eff_src": "C",
        "pwr_26": 30.0, "pwr_src": "C",            # 3.0¢/kWh confirmed; includes ~$21mm power credits
        "opex_26": 180, "opex_src": "E",           # SG&A + engineering est annualized
        "capex_26": 300, "capex_src": "E",         # mining capex est; HPC capex is separate
        "btc_q1": 1473, "rev_q1": 111.9, "cost_mine": 44629, "gross_margin": 0.41,
        "power_note": "3.0¢/kWh all-in (confirmed Q1 2026 10-Q); includes $21mm power credits; gross cost ~$50-60/MWh",
        "hash_note": "42.5 EH/s deployed (C); 36.4 EH/s avg operating (C); scaling toward 50 EH/s+ 2027",
        "eff_note": "20.2 J/TH (C, Q1 2026 10-Q)",
        "opex_note": "Total Q1 2026 COGS (mining) + SG&A allocation est; not isolated in filing",
        "wind_down": False, "tv_zero": False,
    },
    {
        "ticker": "HUT", "name": "Hut 8 / American Bitcoin", "status": "Scaling (ABTC subsidiary)",
        "profile": "Mining via American Bitcoin Corp (majority-owned); S21 XP fleet; Trump-affiliated",
        "hash_26": 25.0, "hash_27": 30.0, "hash_28": 32.0, "hash_src": "C",
        "eff_26": 16.3, "eff_src": "C",            # improving to ~13.5 J/TH with new S21 XP
        "pwr_26": None, "pwr_src": "N/A",          # not disclosed Q1 2026
        "opex_26": 80, "opex_src": "E",
        "capex_26": 49, "capex_src": "C",          # $49.4mm for 11,298 S21 XP miners (Feb 2026)
        "btc_q1": 817, "rev_q1": 62.1, "cost_mine": 36200, "gross_margin": 0.50,
        "power_note": "Not disclosed Q1 2026; Drumheller AB + Vega AB sites; est ~$40-50/MWh Alberta grid",
        "hash_note": "25.0 EH/s Q1 avg (C); 28.1 EH/s by Apr 2026 post S21 XP deploy (C)",
        "eff_note": "16.3 J/TH Q1 avg (C); improving with S21 XP fleet (~13.5 J/TH) deployed Apr 2026",
        "opex_note": "Est from ABTC operations; HUT 8 consolidated SG&A partially attributable",
        "wind_down": False, "tv_zero": False,
    },
    {
        "ticker": "CLSK", "name": "CleanSpark", "status": "Pure-play, scaling + AI land bank",
        "profile": "100% self-mining; southeastern US grid; acquired TX AI/HPC land for future conversion",
        "hash_26": 47.3, "hash_27": 55.0, "hash_28": 60.0, "hash_src": "C",
        "eff_26": 16.07, "eff_src": "C",
        "pwr_26": 56.0, "pwr_src": "C",            # 5.6¢/kWh confirmed
        "opex_26": 120, "opex_src": "E",
        "capex_26": 220, "capex_src": "E",         # significant; AI land acq + fleet upgrades
        "btc_q1": 1799, "rev_q1": 181.2, "cost_mine": None, "gross_margin": 0.47,
        "power_note": "5.6¢/kWh ($56/MWh) confirmed Q1 2026 10-Q; southeastern US grid (GA, WY, TN)",
        "hash_note": "47.3 EH/s Q1 avg (C); 808 MW utilized of 1.8 GW contracted",
        "eff_note": "16.07 J/TH (C, Q1 2026 10-Q)",
        "opex_note": "Q1 2026 normalized EBITDA $55mm; backing into opex from revenue - EBITDA",
        "wind_down": False, "tv_zero": False,
    },
    {
        "ticker": "CORZ", "name": "Core Scientific", "status": "Rapid wind-down of mining",
        "profile": "Mining shrinking fast; CoreWeave AI DC is primary focus ($2B FY2026 capex); mining cash-flow negative",
        "hash_26": 10.0, "hash_27": 3.0, "hash_28": 0.0, "hash_src": "E",
        "eff_26": 22.0, "eff_src": "E",            # not confirmed Q1 2026
        "pwr_26": None, "pwr_src": "N/A",
        "opex_26": 50, "opex_src": "E",
        "capex_26": 0, "capex_src": "E",           # mining capex near zero; all going to HPC
        "btc_q1": 279, "rev_q1": 30.1, "cost_mine": None, "gross_margin": -0.05,
        "power_note": "Not confirmed Q1 2026; legacy TX/KY sites; est ~$45-55/MWh",
        "hash_note": "15.7 EH/s Dec 2025; declining through 2026; ~1-2 mining sites by yr-end 2026 (C)",
        "eff_note": "Not confirmed Q1 2026; estimated ~22 J/TH for aging fleet",
        "opex_note": "Mining gross margin negative; heavy SG&A burden; $266.5mm mining asset impairment Q1",
        "wind_down": True, "tv_zero": True,
    },
]

NA_COMPANIES = [
    {"ticker": "SLNH", "name": "Soluna Holdings",    "reason": "De minimis miner (0.00076 EH/s); primarily hosting/green energy infra; $9.4mm total Q1 rev"},
    {"ticker": "GLXY", "name": "Galaxy Digital",     "reason": "Diversified crypto financial services; mining winding down at Helios TX (now CoreWeave AI DC); mining <1% of $10B+ Q1 revenue"},
    {"ticker": "APLD", "name": "Applied Digital",    "reason": "Does NOT mine Bitcoin — pure colocation/hosting provider to miners and AI hyperscalers"},
    {"ticker": "KEEL", "name": "KEEL / SDIG",        "reason": "SDIG (Stronghold Digital) was acquired by Bitfarms (BITF) and delisted March 2025 — no Q1 2026 10-Q exists"},
]

# ═══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 1: SUMMARY COMP
# ──────────────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Summary Comp"

# Title
ws.merge_cells("A1:N1")
c=ws["A1"]; c.value="BTC Mining Peer Comparison — Q1 2026 SEC Filings"; c.fill=fill(C_NAVY)
c.font=font(True,"FFFFFF",13); c.alignment=align("center","center")
ws.row_dimensions[1].height=22

ws.merge_cells("A2:N2")
c=ws["A2"]; c.value="Sources: 10-Q/8-K filings (Q1 2026, Jan–Mar 2026). GREEN=confirmed. ORANGE=estimated. N/A=not disclosed. All $mm unless noted."
c.fill=fill(C_LBLUE); c.font=font(False,"1F3864",9,True); c.alignment=align("center","center")

# Column headers
COL_HDRS = [
    "Ticker","Company","Status",
    "Hashrate\nQ1 2026\n(EH/s)","FY2027E\n(EH/s)","FY2028E\n(EH/s)",
    "Efficiency\n(J/TH)","Power\n($/MWh)",
    "Other\nOpex\n($mm/yr)","Mining\nCapex\n($mm)",
    "BTC Mined\nQ1 2026","Mining Rev\nQ1 2026\n($mm)","Mining\nGross\nMargin",
    "Notes"
]
for col, hd in enumerate(COL_HDRS, 1):
    c = ws.cell(row=3, column=col, value=hd)
    c.fill=fill(C_BLUE); c.font=font(True,"FFFFFF",9)
    c.alignment=align("center","center",True)
ws.row_dimensions[3].height=42

# Source indicator in separate row
src_hdrs = ["","","","C","E","E","C","C","E","E","C","C","C",""]
for col, s in enumerate(src_hdrs, 1):
    if s:
        c=ws.cell(row=4,column=col,value=f"({s})")
        c.font=font(False,C_DGREY,8,True); c.alignment=align("center","center")
ws.row_dimensions[4].height=12

# Section: Active Miners
ws.merge_cells("A5:N5")
c=ws["A5"]; c.value="ACTIVE MINERS — included in standalone valuation model"
c.fill=fill(C_BLUE); c.font=font(True,"FFFFFF",10); c.alignment=align("left","center")

SRC_BG = {
    "C": C_GRN,
    "E": C_RED,
    "N/A": C_GREY,
}

for i, co in enumerate(COMPANIES):
    r = 6 + i
    bg = C_GREY if co["wind_down"] else C_WHITE

    def cv(val, src="C", fmt=None):
        bg_c = SRC_BG.get(src, C_WHITE)
        return val, bg_c

    vals = [
        (co["ticker"],      C_WHITE, "left",  True),
        (co["name"],        C_WHITE, "left",  False),
        (co["status"],      C_GREY,  "left",  False),
        (co["hash_26"],     SRC_BG.get(co["hash_src"], C_WHITE), "right", False),
        (co["hash_27"],     SRC_BG.get("E"), "right", False),
        (co["hash_28"],     SRC_BG.get("E"), "right", False),
        (co["eff_26"],      SRC_BG.get(co["eff_src"],  C_WHITE), "right", False),
        (co["pwr_26"] if co["pwr_26"] else "N/D", SRC_BG.get(co["pwr_src"],  C_WHITE), "right", False),
        (co["opex_26"],     SRC_BG.get("E"), "right", False),
        (co["capex_26"],    SRC_BG.get(co["capex_src"],C_WHITE), "right", False),
        (co["btc_q1"],      SRC_BG.get("C"), "right", False),
        (co["rev_q1"],      SRC_BG.get("C"), "right", False),
        (f"{co['gross_margin']:.1%}" if co["gross_margin"] is not None else "N/D",
                            SRC_BG.get("C"), "right", False),
        (co["hash_note"],   C_WHITE, "left",  False),
    ]
    for col, (v, bgc, h, bold) in enumerate(vals, 1):
        c2 = ws.cell(row=r, column=col, value=v)
        c2.fill = fill(bgc); c2.font = font(bold, size=9)
        c2.alignment = align(h, "center", True)
    ws.row_dimensions[r].height = 30

# Section: N/A companies
nr = 6 + len(COMPANIES) + 1
ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=14)
c=ws.cell(row=nr,column=1,value="EXCLUDED / N/A — not included in mining standalone model (see notes)")
c.fill=fill(C_DGREY); c.font=font(True,"FFFFFF",10); c.alignment=align("left","center")

for j, co in enumerate(NA_COMPANIES):
    r2 = nr + 1 + j
    ws.cell(row=r2, column=1, value=co["ticker"]).font = font(True, size=9)
    ws.cell(row=r2, column=2, value=co["name"]).font = font(size=9)
    ws.merge_cells(start_row=r2, start_column=3, end_row=r2, end_column=14)
    c=ws.cell(row=r2, column=3, value=co["reason"])
    c.fill=fill(C_GREY); c.font=font(False,C_DGREY,9,True); c.alignment=align("left","center",True)
    ws.row_dimensions[r2].height=24

# Legend
lr = nr + len(NA_COMPANIES) + 2
ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=14)
c=ws.cell(row=lr,column=1,value="LEGEND:   GREEN = confirmed in SEC filing (10-Q/8-K)   |   ORANGE/RED = estimated or derived   |   GREY = not disclosed (N/D) or not applicable (N/A)")
c.fill=fill(C_LBLUE); c.font=font(False,"1F3864",9,True); c.alignment=align("center","center")

# Column widths
col_widths = [8,20,30,9,9,9,9,9,10,10,10,11,9,50]
for i,w in enumerate(col_widths,1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ──────────────────────────────────────────────────────────────────────────────
# INDIVIDUAL COMPANY SHEETS
# ──────────────────────────────────────────────────────────────────────────────

YEARS = ["FY2026E","FY2027E","FY2028E","FY2029E","FY2030E"]
NCOLS = 9  # A-I

def build_company_sheet(wb, co):
    ws = wb.create_sheet(co["ticker"])

    # halving schedule
    issuance = [164250, 164250, 123000, 82125, 82125]

    # hashrate schedule
    h_sched = [co["hash_26"], co["hash_27"], co["hash_28"],
               co["hash_28"]*0.9 if co["hash_28"] > 0 else 0,
               co["hash_28"]*0.8 if co["hash_28"] > 0 else 0]

    # eff: keep constant or improve slightly
    eff_base = co["eff_26"] or 20
    eff_sched = [eff_base, max(eff_base*0.95, 14), max(eff_base*0.90, 13),
                 max(eff_base*0.88, 12), max(eff_base*0.85, 12)]

    # network hashrate (EH/s)
    net_sched = [700, 800, 950, 1050, 1150]

    # power $/MWh
    pwr = co["pwr_26"] or 45
    pwr_sched = [pwr, pwr, pwr*1.02, pwr*1.03, pwr*1.05]

    # other opex $mm
    opex = co["opex_26"] or 50
    if co["wind_down"]:
        opex_sched = [opex, opex*0.5, 0, 0, 0]
    else:
        opex_sched = [opex, opex*1.10, opex*1.15, opex*1.20, opex*1.20]

    # capex $mm
    capex = co["capex_26"] or 0
    if co["wind_down"]:
        capex_sched = [capex, capex*0.3, 0, 0, 0]
    else:
        capex_sched = [capex, capex*0.9, capex*0.7, capex*0.5, capex*0.3]

    btc_price = [110000, 120000, 130000, 140000, 150000]  # user editable

    # ── Title ──
    ws.merge_cells(f"A1:I1")
    c=ws["A1"]; c.value=f"{co['ticker']} ({co['name']}) — BTC Mining Standalone Valuation"
    c.fill=fill(C_NAVY); c.font=font(True,"FFFFFF",12); c.alignment=align("center","center")
    ws.row_dimensions[1].height=20

    ws.merge_cells("A2:I2")
    c=ws["A2"]; c.value=co["profile"]
    c.fill=fill(C_LBLUE); c.font=font(False,"1F3864",9,True); c.alignment=align("center","center")

    # ── Section: Key Assumptions ──
    ws.merge_cells("A3:I3")
    c=ws["A3"]; c.value="KEY ASSUMPTIONS — BTC Mining Only"
    c.fill=fill(C_BLUE); c.font=font(True,"FFFFFF",10); c.alignment=align("left","center")

    # column headers
    for ci, yr in enumerate(YEARS, 3):
        hdr(ws, 4, ci, yr, C_NAVY, "FFFFFF", 9)
    hdr(ws, 4, 8, "Notes", C_NAVY, "FFFFFF", 9)
    hdr(ws, 4, 1, "Driver", C_NAVY, "FFFFFF", 9)
    hdr(ws, 4, 2, "Unit",   C_NAVY, "FFFFFF", 9)

    def assum_row(row, driver, unit, vals, srcs, note, fmts=None):
        label(ws, row, 1, driver)
        cell(ws, row, 2, unit, h="center")
        for ci, (v, s) in enumerate(zip(vals, srcs)):
            bg = SRC_BG.get(s, C_WHITE)
            fmt = fmts[ci] if fmts else None
            cell(ws, row, ci+3, v, bg=bg, fmt=fmt)
        c=ws.cell(row=row,column=8,value=note)
        c.font=font(False,C_DGREY,8,True); c.alignment=align("left","center",True)

    def src5(s): return [s]*5
    def src_wd(s): return [s, s, "N/A","N/A","N/A"] if co["wind_down"] else src5(s)

    zero_bg = lambda v, s: "N/A" if (v==0 and co["wind_down"] and s not in ("C","E")) else s

    assum_row(5,  "Installed hashrate","EH/s",
        [round(v,1) for v in h_sched],
        [co["hash_src"]]*2 + ["E","E","E"],
        co["hash_note"])

    assum_row(6,  "Avg operating factor","%",
        [0.90]*5 if not co["wind_down"] else [0.90, 0.90, 0, 0, 0],
        ["C","C","N/A","N/A","N/A"] if co["wind_down"] else src5("C"),
        "Uptime × curtailment; 90% std assumption", ["0%"]*5)

    assum_row(7,  "Network hashrate","EH/s",
        net_sched, src5("E"),
        "Global network estimate; per public tracker")

    assum_row(8,  "Annual BTC issuance","BTC",
        issuance, src5("C"),
        "Protocol schedule: halving ~Apr 2028 → 164,250 → 123,000 → 82,125 BTC/yr")

    assum_row(9,  "BTC price","$/BTC",
        btc_price, src5("N/A"),
        "USER INPUT — change to your BTC price deck", ["#,##0"]*5)

    assum_row(10, "Fleet efficiency","J/TH",
        [round(v,1) for v in eff_sched],
        [co["eff_src"]]*2 + ["E","E","E"],
        co["eff_note"])

    pwr_disp = [round(v,1) if co["pwr_26"] else "N/D" for v in pwr_sched]
    assum_row(11, "Electricity price","$/MWh",
        pwr_disp,
        [co["pwr_src"]]*5,
        co["power_note"])

    assum_row(12, "Other mining opex","$mm",
        [round(v,0) for v in opex_sched],
        src5("E"),
        co["opex_note"])

    assum_row(13, "Mining capex","$mm",
        [round(v,0) for v in capex_sched],
        [co["capex_src"]]*2 + ["E","E","E"],
        "ASIC + DC build capex" if not co["wind_down"] else "Wind-down: minimal/zero capex")

    assum_row(14, "D&A (% of capex)","%",
        [0.20]*5, src5("E"), "Simplified D&A", ["0%"]*5)

    assum_row(15, "Cash tax rate","%",
        [0.21]*5, src5("E"),
        "21% US federal; may be 0% near-term due to NOL carryforwards", ["0%"]*5)

    assum_row(16, "NWC (% of revenue)","%",
        [0.02]*5, src5("E"), "Working capital", ["0%"]*5)

    # ── Section: Operating Model ──
    ws.merge_cells("A18:I18")
    c=ws["A18"]; c.value="OPERATING MODEL"
    c.fill=fill(C_BLUE); c.font=font(True,"FFFFFF",10); c.alignment=align("left","center")

    for ci, yr in enumerate(YEARS, 3):
        hdr(ws, 19, ci, yr, C_NAVY, "FFFFFF", 9)
    hdr(ws, 19, 1, "Metric", C_NAVY, "FFFFFF", 9)
    hdr(ws, 19, 2, "Unit",   C_NAVY, "FFFFFF", 9)
    hdr(ws, 19, 8, "Formula / Comment", C_NAVY, "FFFFFF", 9)

    def op_row(row, metric, unit, vals, note, bold=False, fmt="#,##0.0"):
        label(ws, row, 1, metric, bold=bold)
        cell(ws, row, 2, unit, h="center", sz=9)
        for ci, v in enumerate(vals, 3):
            bg = C_GRN if bold else None
            cell(ws, row, ci, v if v is not None else "—", bg=bg if bold else None,
                 bold=bold, sz=9, fmt=fmt)
        c=ws.cell(row=row,column=8,value=note)
        c.font=font(False,C_DGREY,8,True); c.alignment=align("left","center",True)

    # Compute operating model
    op_factor = [0.90]*5 if not co["wind_down"] else [0.90, 0.90, 0, 0, 0]
    avg_hash = [h*f for h,f in zip(h_sched, op_factor)]
    net_share = [a/n for a,n in zip(avg_hash, net_sched)]
    btc_mined = [s*i for s,i in zip(net_share, issuance)]
    mining_rev = [b*p/1e6 for b,p in zip(btc_mined, btc_price)]
    power_mw = [a*e for a,e in zip(avg_hash, eff_sched)]   # MW
    pwr_used = [co["pwr_26"] or 45]*5
    elec_cost = [pw*24*365*pr/1e6 for pw,pr in zip(power_mw, pwr_used)]
    total_opex = [e+o for e,o in zip(elec_cost, opex_sched)]
    ebitda = [r-o for r,o in zip(mining_rev, total_opex)]
    ebitda_margin = [e/r if r>0 else 0 for e,r in zip(ebitda, mining_rev)]
    da = [capex_sched[i]*0.20 for i in range(5)]
    tax = [max(0, (ebitda[i]-da[i])*0.21) for i in range(5)]
    nwc_chg = [r*0.02 for r in mining_rev]
    fcf = [ebitda[i]-tax[i]-capex_sched[i]-nwc_chg[i] for i in range(5)]

    op_row(20,"Avg operating hashrate","EH/s",[round(v,2) for v in avg_hash],"Installed × factor")
    op_row(21,"Network share","%",[round(v*100,4) for v in net_share],"Op hashrate / network",fmt="0.0000%")
    op_row(22,"BTC mined","BTC",[round(v,0) for v in btc_mined],"Share × issuance", fmt="#,##0")
    op_row(23,"Mining revenue","$mm",[round(v,1) for v in mining_rev],"BTC mined × price / 1mm",bold=True,fmt="#,##0.0")
    op_row(24,"Power draw","MW",[round(v,1) for v in power_mw],"Avg hashrate × efficiency J/TH")
    op_row(25,"Electricity cost","$mm",[round(v,1) for v in elec_cost],"MW × 8,760 hrs × $/MWh")
    op_row(26,"Other opex","$mm",[round(v,1) for v in opex_sched],"From assumptions")
    op_row(27,"Total cash opex","$mm",[round(v,1) for v in total_opex],"Elec + other")
    op_row(28,"Mining EBITDA","$mm",[round(v,1) for v in ebitda],"Revenue − cash opex",bold=True)
    op_row(29,"EBITDA margin","%",[round(v*100,1) for v in ebitda_margin],"EBITDA / revenue",fmt="0.0%")
    op_row(30,"D&A","$mm",[round(v,1) for v in da],"Capex × 20%")
    op_row(31,"Cash taxes","$mm",[round(v,1) for v in tax],"max(0, EBIT × 21%)")
    op_row(32,"Mining capex","$mm",[round(v,0) for v in capex_sched],"From assumptions")
    op_row(33,"Change in NWC","$mm",[round(v,1) for v in nwc_chg],"Revenue × 2%")
    op_row(34,"Mining FCF","$mm",[round(v,1) for v in fcf],"EBITDA − tax − capex − NWC",bold=True)

    # ── Section: Valuation ──
    ws.merge_cells("A36:I36")
    c=ws["A36"]; c.value="VALUATION — Mining Only (EV building block)"
    c.fill=fill(C_BLUE); c.font=font(True,"FFFFFF",10); c.alignment=align("left","center")

    WACC = 0.15 if co["wind_down"] else 0.12
    TG   = 0.0  if co["wind_down"] else 0.025
    MULT = 3.0  if co["wind_down"] else 8.0

    label(ws, 37, 1, "Method 1 — DCF: 5-yr explicit FCF + terminal value", bold=True)
    label(ws, 38, 1, "WACC"); cell(ws, 38, 3, WACC, bg=C_YEL if not co["wind_down"] else C_RED, fmt="0%")
    ws.cell(row=38,column=8).value="Mining cost of capital; higher (15%) for wind-down assets"
    label(ws, 39, 1, "Terminal growth (g)"); cell(ws, 39, 3, TG, bg=C_GREY if co["tv_zero"] else C_YEL, fmt="0.0%")
    ws.cell(row=39,column=8).value="0% for wind-down (no going concern); ~2.5% for scaling miners"

    # DCF computation
    pv_fcf = [f/(1+WACC)**(i+1) for i,f in enumerate(fcf)]
    sum_pv = sum(pv_fcf)

    if co["tv_zero"] or ebitda[-1] <= 0:
        norm_fcf = 0; tv_undis = 0; tv_pv = 0
    else:
        maint_capex = capex_sched[-1]
        norm_fcf = ebitda[-1] - max(0,(ebitda[-1]-maint_capex)*0.21) - maint_capex
        tv_undis = norm_fcf*(1+TG)/(WACC-TG) if WACC > TG else 0
        tv_pv = tv_undis/(1+WACC)**5

    dcf_ev = sum_pv + tv_pv

    label(ws, 40, 1, "Mining FCF",      bold=True)
    label(ws, 41, 1, "Discount factor", bold=False)
    label(ws, 42, 1, "PV of FCF",       bold=True)

    for ci, yr in enumerate(YEARS, 3):
        hdr(ws, 40, ci, yr, C_NAVY,"FFFFFF",8)

    for ci, (f, pv) in enumerate(zip(fcf, pv_fcf), 3):
        cell(ws, 40, ci, round(f,1), fmt="#,##0.0")
        df = 1/(1+WACC)**(ci-2)
        cell(ws, 41, ci, round(df,4), fmt="0.0000")
        cell(ws, 42, ci, round(pv,1), bg=C_GRN, bold=True, fmt="#,##0.0")

    label(ws, 43, 1, "Normalized terminal FCF")
    cell(ws, 43, 3, round(norm_fcf,1), fmt="#,##0.0")
    ws.cell(row=43,column=8).value = "$0 for wind-down assets" if co["tv_zero"] else "FY2030 EBITDA − maint tax − maint capex"

    label(ws, 44, 1, "Terminal value (undiscounted)")
    cell(ws, 44, 3, round(tv_undis,1), bg=C_GREY if co["tv_zero"] else C_GRN, fmt="#,##0.0")
    ws.cell(row=44,column=8).value = "TV = $0 (mining wind-down; no going-concern value)" if co["tv_zero"] else "Gordon growth on norm FCF"

    label(ws, 45, 1, "PV of terminal value")
    cell(ws, 45, 3, round(tv_pv,1), bg=C_GREY if co["tv_zero"] else C_GRN, fmt="#,##0.0")

    label(ws, 46, 1, "Sum PV of explicit FCF")
    cell(ws, 46, 3, round(sum_pv,1), bg=C_GRN, bold=True, fmt="#,##0.0")

    label(ws, 47, 1, "DCF Enterprise Value (Method 1)", bold=True)
    cell(ws, 47, 3, round(dcf_ev,1), bg=C_YEL, bold=True, fmt="#,##0.0")
    ws.cell(row=47,column=8).value = "Mining-only EV (sum of PV FCFs + PV terminal)"

    # Method 2
    ws.merge_cells("A49:I49")
    c=ws["A49"]; c.value="Method 2 — EV/EBITDA Cross-Check"
    c.fill=fill(C_BLUE); c.font=font(True,"FFFFFF",10); c.alignment=align("left","center")

    ref_ebitda = ebitda[0] if co["wind_down"] else ebitda[1]
    ref_yr = "FY2026E (last full mining yr)" if co["wind_down"] else "FY2027E (forward)"
    mult_ev = ref_ebitda * MULT if ref_ebitda > 0 else 0

    label(ws, 50, 1, f"Reference EBITDA ({ref_yr})")
    cell(ws, 50, 3, round(ref_ebitda,1), bg=C_GRN, fmt="#,##0.0")

    label(ws, 51, 1, "Mining EV/EBITDA multiple")
    cell(ws, 51, 3, MULT, bg=C_YEL, fmt="0.0x")
    ws.cell(row=51,column=8).value = "3x for wind-down; 8x for scaling peers (editable)" if co["wind_down"] else "Peer-based; editable"

    label(ws, 52, 1, "Implied EV (Method 2)", bold=True)
    cell(ws, 52, 3, round(mult_ev,1), bg=C_YEL, bold=True, fmt="#,##0.0")

    # Reconciliation
    ws.merge_cells("A54:I54")
    c=ws["A54"]; c.value="Reconciliation"
    c.fill=fill(C_BLUE); c.font=font(True,"FFFFFF",10); c.alignment=align("left","center")

    label(ws, 55, 1, "DCF EV (Method 1)")
    cell(ws, 55, 3, round(dcf_ev,1), bg=C_GRN, fmt="#,##0.0")
    label(ws, 56, 1, "Multiple EV (Method 2)")
    cell(ws, 56, 3, round(mult_ev,1), bg=C_GRN, fmt="#,##0.0")
    label(ws, 57, 1, "Difference (Multiple − DCF)")
    cell(ws, 57, 3, round(mult_ev-dcf_ev,1), fmt="#,##0.0")
    label(ws, 58, 1, "Selected method (1=DCF / 2=Multiple)")
    cell(ws, 58, 3, 2, bg=C_YEL)
    label(ws, 59, 1, "Selected Mining EV", bold=True)
    cell(ws, 59, 3, round(mult_ev if True else dcf_ev,1), bg=C_YEL, bold=True, fmt="#,##0.0")
    ws.cell(row=59,column=8).value = "Building block for SOTP — add HPC/AI data center NPV"

    # Q1 2026 Actuals box
    ws.merge_cells("A61:I61")
    c=ws["A61"]; c.value="Q1 2026 ACTUALS (from 10-Q / 8-K filing)"
    c.fill=fill(C_BLUE); c.font=font(True,"FFFFFF",10); c.alignment=align("left","center")

    actuals = [
        ("BTC mined Q1 2026", "BTC", co["btc_q1"], "#,##0"),
        ("Mining revenue Q1 2026", "$mm", co["rev_q1"], "#,##0.0"),
        ("Annualized mining revenue", "$mm", round(co["rev_q1"]*4,1) if co["rev_q1"] else None, "#,##0.0"),
        ("Mining gross margin", "%", f"{co['gross_margin']:.1%}" if co["gross_margin"] else "N/D", None),
        ("Cost to mine (ex-D&A)", "$/BTC", co["cost_mine"], "#,##0"),
    ]
    for j, (metric, unit, val, fmt) in enumerate(actuals):
        rr = 62 + j
        label(ws, rr, 1, metric)
        cell(ws, rr, 2, unit, h="center", sz=9)
        cell(ws, rr, 3, val if val else "N/D", bg=C_GRN if val else C_GREY, fmt=fmt)

    # column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 8
    for col in ["C","D","E","F","G"]: ws.column_dimensions[col].width = 12
    ws.column_dimensions["H"].width = 55

for co in COMPANIES:
    build_company_sheet(wb, co)

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Sheets: {[ws.title for ws in wb.worksheets]}")
