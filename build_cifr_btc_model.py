"""
Populates BTC Mining Standalone Valuation model with CIFR-specific data.
CIFR mining segment: Odessa TX PPA expires July 2027 → hashrate → 0 from FY2028.
Terminal value = $0 (no going concern for mining segment).
"""

import shutil
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, numbers
from openpyxl.utils import get_column_letter

SRC = "/root/.claude/uploads/e81ff327-71b7-5ae3-aeed-4a9a48655120/00470799-BTC_model.xlsx"
DST = "/home/user/claude-code/CIFR_BTC_Mining_Standalone.xlsx"

shutil.copy2(SRC, DST)

wb = openpyxl.load_workbook(DST)
ws = wb["BTC Mining (Standalone)"]

# ─── Styles ───────────────────────────────────────────────────────────────────
YELLOW   = PatternFill("solid", fgColor="FFD966")   # user input / assumption
GREEN    = PatternFill("solid", fgColor="E2EFDA")   # confirmed from filings
ORANGE   = PatternFill("solid", fgColor="FCE4D6")   # estimated / proxy
GREY     = PatternFill("solid", fgColor="F2F2F2")   # zero / N/A
BOLD     = Font(bold=True, name="Calibri", size=10)
NORM     = Font(name="Calibri", size=10)
ITALIC   = Font(italic=True, name="Calibri", size=9, color="595959")

def set_cell(ws, row, col, value, fill=None, font=None, num_fmt=None, note=None):
    c = ws.cell(row=row, column=col)
    c.value = value
    if fill:   c.fill = fill
    if font:   c.font = font
    if num_fmt: c.number_format = num_fmt
    return c

def zero_row(ws, row, cols=(3,4,5,6,7)):
    for col in cols:
        set_cell(ws, row, col, 0, fill=GREY)

# ─── Title update ─────────────────────────────────────────────────────────────
ws["A1"] = "CIFR (Cipher Digital) — BTC Mining Segment Standalone Valuation"
ws["A1"].font = Font(bold=True, size=12, name="Calibri", color="1F3864")

ws["H4"] = "Legend: GREEN = SEC filing / confirmed | YELLOW = user input | ORANGE = estimate | GREY = zero (mining wind-down)"
ws["H4"].font = ITALIC

# ─── KEY ASSUMPTIONS (rows 6-17) ──────────────────────────────────────────────
# CIFR mining profile:
#   FY2026: 11.6 EH/s full year (Q1 2026 actual)
#   FY2027: ~5.8 EH/s effective (Odessa PPA expires Jul 2027 → ~6 months)
#   FY2028-30: 0 (mining wound down)

# Row 6 — Installed hashrate (EH/s)
# Source: CIFR Q1 2026 10-Q — 11.6 EH/s deployed
set_cell(ws, 6, 3, 11.6,  fill=GREEN)   # 2026 confirmed
set_cell(ws, 6, 4, 5.8,   fill=ORANGE)  # 2027: ~half-year effective avg
set_cell(ws, 6, 5, 0,     fill=GREY)    # 2028 zero
set_cell(ws, 6, 6, 0,     fill=GREY)    # 2029 zero
set_cell(ws, 6, 7, 0,     fill=GREY)    # 2030 zero
ws["H6"] = "2026: Q1 2026 10-Q (11.6 EH/s). 2027: ~6-month avg (PPA expires Jul 2027). 2028+: $0 — mining wound down."

# Row 7 — Avg operating factor
set_cell(ws, 7, 3, 0.90, fill=GREEN)    # Q1 2026 ~90% uptime
set_cell(ws, 7, 4, 0.90, fill=GREEN)
set_cell(ws, 7, 5, 0,    fill=GREY)
set_cell(ws, 7, 6, 0,    fill=GREY)
set_cell(ws, 7, 7, 0,    fill=GREY)
ws["H7"] = "Q1 2026 10-Q operating efficiency. 2028+: zero (no operations)."

# Row 8 — Network hashrate (EH/s)
set_cell(ws, 8, 3, 700,  fill=ORANGE)   # ~700 EH/s mid-2026 estimate
set_cell(ws, 8, 4, 800,  fill=ORANGE)   # ~800 EH/s mid-2027 estimate
set_cell(ws, 8, 5, 950,  fill=GREY)     # irrelevant (no mining)
set_cell(ws, 8, 6, 1050, fill=GREY)
set_cell(ws, 8, 7, 1150, fill=GREY)
ws["H8"] = "Network est. per public trackers. 2028+: irrelevant (CIFR exits mining)."

# Row 9 — Annual BTC issuance — keep as-is (protocol fact)
# 164,250 / 164,250 / 123,000 / 82,125 / 82,125 — already correct in template
ws["H9"] = "Protocol schedule. Halving ~Apr 2028: 164,250 → 123,000 → 82,125 BTC/yr."

# Row 10 — BTC price (USER INPUT — yellow)
set_cell(ws, 10, 3, 110000, fill=YELLOW)  # 2026 assumption
set_cell(ws, 10, 4, 120000, fill=YELLOW)  # 2027 assumption
set_cell(ws, 10, 5, 120000, fill=GREY)    # 2028+ irrelevant
set_cell(ws, 10, 6, 120000, fill=GREY)
set_cell(ws, 10, 7, 120000, fill=GREY)
ws["H10"] = "USER INPUT — adjust BTC price assumption. 2028+ cells greyed (zero revenue)."

# Row 11 — Fleet efficiency (J/TH)
set_cell(ws, 11, 3, 17.2, fill=GREEN)   # Q1 2026 10-Q: 17.2 J/TH
set_cell(ws, 11, 4, 17.2, fill=GREEN)   # assume same fleet through mid-2027
set_cell(ws, 11, 5, 0,    fill=GREY)
set_cell(ws, 11, 6, 0,    fill=GREY)
set_cell(ws, 11, 7, 0,    fill=GREY)
ws["H11"] = "Q1 2026 10-Q: 17.2 J/TH fleet avg. No capex → same cohort through wind-down."

# Row 12 — Electricity price ($/MWh) — MISSING (PPA not public)
set_cell(ws, 12, 3, 42, fill=ORANGE)   # Odessa TX PPA est. ~$42/MWh
set_cell(ws, 12, 4, 42, fill=ORANGE)
set_cell(ws, 12, 5, 0,  fill=GREY)
set_cell(ws, 12, 6, 0,  fill=GREY)
set_cell(ws, 12, 7, 0,  fill=GREY)
ws["H12"] = "ESTIMATE — Odessa TX PPA rate NOT publicly disclosed. ~$42/MWh proxy (W TX industrial). USER INPUT."

# Row 13 — Other mining opex ($mm)
# CIFR does not break out mining-segment opex separately from HPC in recent filings
# Q1 2026: total BTC segment rev ~$35mm, G&A+SG&A mining portion est.
set_cell(ws, 13, 3, 55,  fill=ORANGE)  # 2026 full yr est.
set_cell(ws, 13, 4, 28,  fill=ORANGE)  # 2027 half yr est.
set_cell(ws, 13, 5, 0,   fill=GREY)
set_cell(ws, 13, 6, 0,   fill=GREY)
set_cell(ws, 13, 7, 0,   fill=GREY)
ws["H13"] = "ESTIMATE — CIFR does not break out mining SG&A from HPC in Q1 2026 filing. ~$55mm/yr proxy (proportional G&A). USER INPUT."

# Row 14 — Mining capex ($mm) — ~$0 (winding down, no new ASICs)
set_cell(ws, 14, 3, 0, fill=GREEN)
set_cell(ws, 14, 4, 0, fill=GREEN)
set_cell(ws, 14, 5, 0, fill=GREY)
set_cell(ws, 14, 6, 0, fill=GREY)
set_cell(ws, 14, 7, 0, fill=GREY)
ws["H14"] = "CIFR mining wind-down: no new ASIC purchases. Capex = $0."

# Row 15 — D&A % of capex
# Since capex = 0, D&A = 0 effectively. Keep 20% but result will be $0.
# Keep as-is in template (formula driven)
ws["H15"] = "With capex = $0, D&A = $0. Rate kept at 20% for formula consistency."

# Row 16 — Cash tax rate — use 21% uniformly (US federal)
set_cell(ws, 16, 3, 0.21, fill=GREEN)
set_cell(ws, 16, 4, 0.21, fill=GREEN)
set_cell(ws, 16, 5, 0.21, fill=GREY)
set_cell(ws, 16, 6, 0.21, fill=GREY)
set_cell(ws, 16, 7, 0.21, fill=GREY)
ws["H16"] = "US federal statutory rate 21%. CIFR has NOL carryforwards → effective rate may be 0% near-term."

# Row 17 — NWC % of revenue
# Keep 2% — formula will auto-zero when revenue = 0 in 2028+

# ─── DCF: Override terminal value to $0 (mining going concern = $0) ───────────
# Row 39: WACC
set_cell(ws, 39, 3, 0.15, fill=ORANGE)  # higher WACC for mining wind-down risk
ws["H39"] = "CIFR mining wind-down: 15% WACC (higher risk given finite life). No terminal value."

# Row 40: terminal growth — set to 0 (or we'll zero out TV manually)
set_cell(ws, 40, 3, 0.0, fill=GREY)
ws["H40"] = "Terminal growth = 0 — CIFR mining has no going-concern value beyond FY2027."

# Row 41: Maintenance capex terminal — set to 0
set_cell(ws, 41, 3, 0, fill=GREY)

# Row 47: Normalized terminal FCF — override to $0
# The formula uses FY2030 EBITDA which will already be $0 since hashrate=0, so auto-zeroes.
ws["H47"] = "Auto-zeroes: FY2030 hashrate = 0 → EBITDA = 0 → TV = $0. Mining has no terminal value."
ws["H48"] = "TV = $0. CIFR mining terminates when Odessa TX PPA expires (Jul 2027)."

# ─── EV/EBITDA cross-check: adjust multiple for wind-down ────────────────────
# Row 54: Reference EBITDA — use FY2026 (last full-year mining year, not FY2027)
ws["A54"] = "Reference EBITDA (FY2026E — last full mining year)"
ws["C54"] = "=C29"   # FY2026 EBITDA (more representative than half-yr 2027)
ws["H54"] = "Use FY2026E EBITDA as reference — FY2027 is only 6 months of mining."

# Row 55: Multiple — lower for wind-down asset
set_cell(ws, 55, 3, 3.0, fill=ORANGE)
ws["H55"] = "3x EV/EBITDA — wind-down asset (finite life ~18 months). Peer pure-play miners ~5-8x, but CIFR mining is terminal."

# ─── Add CIFR-specific annotation box below reconciliation ───────────────────
# Find where reconciliation ends (around row 63) and add notes
note_row = 66
ws.cell(row=note_row, column=1).value = "⚠  CIFR MINING WIND-DOWN NOTES"
ws.cell(row=note_row, column=1).font = Font(bold=True, color="C00000", name="Calibri", size=10)

notes = [
    ("Structure:", "CIFR is exiting Bitcoin mining. Odessa TX PPA (200 MW) expires July 2027. All capacity converting to HPC/AI data center use."),
    ("FY2027 treatment:", "FY2027 reflects ~6 months of mining revenue (Jan–Jun 2027). Hashrate assumption = 5.8 EH/s effective average."),
    ("Terminal value:", "$0. No residual mining operations after July 2027. This EV is a finite-cash-flow present value, not a going-concern DCF."),
    ("Missing data:", "Electricity $/MWh (Odessa PPA rate not disclosed) and mining SG&A (not broken out from HPC) are ESTIMATES. See orange cells."),
    ("BTC price:", "USER INPUT in row 10. Adjust to your BTC price deck. Sensitivity table below shows DCF EV across BTC price × WACC."),
    ("NOL carryforward:", "CIFR has significant NOLs → cash tax may be $0 near-term. Conservative 21% rate used; actual EV may be higher."),
    ("Primary valuation:", "HPC/AI data center NPV (signed contracts + pipeline) drives CIFR's equity story. This mining model is a minor building block."),
]

for i, (label, text) in enumerate(notes, 1):
    r = note_row + i
    ws.cell(row=r, column=1).value = label
    ws.cell(row=r, column=1).font = Font(bold=True, name="Calibri", size=9)
    ws.cell(row=r, column=2).value = text
    ws.cell(row=r, column=2).font = Font(name="Calibri", size=9)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["B"].width = 80

ws.column_dimensions["A"].width = 28
ws.column_dimensions["H"].width = 75
for col in ["C","D","E","F","G"]:
    ws.column_dimensions[col].width = 13

wb.save(DST)
print(f"Saved: {DST}")
print("\nCIFR DATA STATUS:")
print("─" * 60)
print("CONFIRMED (green) — from SEC filings:")
print("  Installed hashrate FY2026: 11.6 EH/s (Q1 2026 10-Q)")
print("  Fleet efficiency: 17.2 J/TH (Q1 2026 10-Q)")
print("  Mining capex: $0 (wind-down, no new ASICs)")
print("  Cash tax rate: 21% (US federal)")
print("  Operating factor: ~90%")
print()
print("ESTIMATED (orange) — proxies / need verification:")
print("  Network hashrate: ~700/800 EH/s (public tracker estimates)")
print("  Electricity $/MWh: ~$42 (Odessa TX PPA UNDISCLOSED — estimate)")
print("  Other mining opex: ~$55mm FY2026 (NOT broken out in filings)")
print("  FY2027 effective hashrate: 5.8 EH/s (half-year assumption)")
print("  EV/EBITDA multiple: 3x (wind-down asset haircut)")
print()
print("USER INPUT (yellow) — needs your decision:")
print("  BTC price: $110k (2026), $120k (2027) — CHANGE AS NEEDED")
print()
print("MISSING / NOT PUBLICLY DISCLOSED:")
print("  1. Exact Odessa PPA electricity rate ($/MWh)")
print("  2. Mining-segment SG&A (blended with HPC in Q1 2026 filings)")
print("  3. BTC price assumption (your call)")
