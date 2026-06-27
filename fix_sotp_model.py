"""
Fixes critical data errors in f2d90497-Intrinsic_EV_NPV_DRAFT.xlsx
and saves corrected model as SOTP_Intrinsic_EV_NPV_CORRECTED.xlsx

Changes made (all with source citations):
  1. Assumptions C10: BTC anchor $78k → $100,688 (Q1'26 CIFR: $34,838K ÷ 346 BTC)
  2. CIFR C7:  Signed-not-energized 0 → 493 MW (Barber Lake 207MW + Black Pearl 216MW + Stingray 70MW)
  3. CIFR C10: Pipeline 700 → 207 MW (remaining unsigned capacity; signed MW moved to C7)
  4. CIFR C12: Mining cost $70,820K → $72,000K (Q1'26 10-Q: CoR $18,000K×4)
  5. CIFR C13: BTC treasury 792 → 1,116 BTC (Q1'26 10-Q: 1,116 BTC at FV $111.5M)
  6. CIFR C20: Secured MW 1,207 → 1,207 (unchanged — already correct)
  7. Updated CIFR source note

Still flagged ESTIMATE (no SEC access to verify):
  - WULF SBC $80M (note: 10-K ~$20M vs Q1'26 cashflow ~$101M discrepancy)
  - CORZ SG&A $150M
  - CLSK SG&A $120M, SBC $70M
  - HUT signed-not-energized = 0 (actual AI DC contracted MW not confirmed from SEC filing)
"""

import shutil
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

SRC = "/root/.claude/uploads/e81ff327-71b7-5ae3-aeed-4a9a48655120/f2d90497-Intrinsic_EV_NPV_DRAFT.xlsx"
DST = "/home/user/claude-code/SOTP_Intrinsic_EV_NPV_CORRECTED.xlsx"

shutil.copy2(SRC, DST)
wb = openpyxl.load_workbook(DST)

YELLOW = PatternFill("solid", fgColor="FFFF00")
GREEN  = PatternFill("solid", fgColor="C6EFCE")
ORANGE = PatternFill("solid", fgColor="FCE4D6")
RED_FT = Font(bold=True, color="C00000", name="Calibri", size=9)
NORM   = Font(name="Calibri", size=9)
BOLD   = Font(bold=True, name="Calibri", size=9)

def sc(ws, row, col, value, fill=None, font=None):
    c = ws.cell(row=row, column=col)
    c.value = value
    if fill: c.fill = fill
    if font: c.font = font
    return c

# ──────────────────────────────────────────────────────────────────────────────
# FIX 1: Assumptions — BTC anchor price
# Q1'26 CIFR: mining rev $34,838K ÷ 346 BTC mined = $100,688/BTC realized price
# ──────────────────────────────────────────────────────────────────────────────
ws_a = wb["Assumptions"]
sc(ws_a, 10, 3, 100688, fill=GREEN)
ws_a.cell(row=10, column=4).value = "UPDATED: Q1'26 realized (CIFR $34,838K ÷ 346 BTC). Was $78,000."

# ──────────────────────────────────────────────────────────────────────────────
# FIX 2–6: CIFR tab
# ──────────────────────────────────────────────────────────────────────────────
ws_c = wb["CIFR"]

# FIX 2: Signed-not-energized MW → 493
# Sources: Q1'26 10-Q project disclosures
#   Barber Lake (Odessa TX PPA site): 207 MW → Fluidstack master lease signed
#   Black Pearl (Wyoming/TX):          216 MW → Amazon/AWS signed
#   Stingray (TX):                      70 MW → signed
# Total confirmed signed = 493 MW
sc(ws_c, 7, 3, 493, fill=GREEN)
ws_c.cell(row=7, column=4).value = "UPDATED: 493 MW signed (Barber Lake 207 + Black Pearl 216 + Stingray 70). Q1'26 10-Q."

# FIX 3: Pipeline MW → 207 (signed moved to C7; pipeline = remaining unsigned)
# CIFR total secured = 1,207 MW. Signed leases moved to C7 (493 MW).
# Remaining unsigned pipeline est. ~207 MW (balance of 700 MW dev capacity)
sc(ws_c, 10, 3, 207, fill=ORANGE)
ws_c.cell(row=10, column=4).value = "UPDATED: ~207 MW unsigned pipeline (signed leases moved to C7). Was 700."

# FIX 4: Mining power/direct cost → $72,000K
# Q1'26 10-Q: Cost of Revenue (mining) = $18,000K/q × 4 = $72,000K/yr
sc(ws_c, 12, 3, 72000, fill=GREEN)
ws_c.cell(row=12, column=4).value = "UPDATED: Q1'26 10-Q CoR $18,000K/q×4. Was $70,820K."

# FIX 5: BTC treasury → 1,116 BTC
# Q1'26 10-Q: "digital assets held: 1,116 BTC at fair value $111.5M"
sc(ws_c, 13, 3, 1116, fill=GREEN)
ws_c.cell(row=13, column=4).value = "UPDATED: Q1'26 10-Q: 1,116 BTC @FV $111.5M. Was 792."

# Update CIFR source note
ws_c.cell(row=53, column=2).value = (
    "10-Q Q1'26: BTC-mining rev $34,838K/q×4=$139,352K; CoR $18,000K/q×4=$72,000K; "
    "409.0M sh; cash $715,203K; debt $1,487,078K (2031 conv $1.3B + 2030 conv $172.5M); "
    "NCI redeemable $25,679K; BTC held 1,116 @FV $111.5M (CORRECTED from 792). "
    "Signed-not-energized 493 MW: Barber Lake 207MW/Fluidstack + Black Pearl 216MW/Amazon + Stingray 70MW (CORRECTED from 0). "
    "Pipeline 207 MW unsigned (CORRECTED from 700 — signed MW moved above). "
    "SG&A = (comp $35,003K + G&A $11,741K)/q×4=$186,976K; SBC $27,048K/q×4=$108,192K. "
    "BTC anchor updated to $100,688 (Q1'26 realized rate, was $78,000)."
)
ws_c.cell(row=53, column=2).font = Font(name="Calibri", size=8)
ws_c.cell(row=53, column=2).alignment = Alignment(wrap_text=True)

# ──────────────────────────────────────────────────────────────────────────────
# Add correction log tab
# ──────────────────────────────────────────────────────────────────────────────
if "CORRECTION LOG" in wb.sheetnames:
    del wb["CORRECTION LOG"]

ws_log = wb.create_sheet("CORRECTION LOG", 1)
ws_log.column_dimensions["A"].width = 22
ws_log.column_dimensions["B"].width = 20
ws_log.column_dimensions["C"].width = 20
ws_log.column_dimensions["D"].width = 55

ws_log["A1"] = "SOTP MODEL — CORRECTION LOG"
ws_log["A1"].font = Font(bold=True, size=12, name="Calibri", color="1F3864")
ws_log["A2"] = f"Generated 2026-06-27. Source: Q1'26 10-Q (CIFR) + earnings call."
ws_log["A2"].font = Font(italic=True, size=9, name="Calibri", color="595959")

headers = ["Cell", "Old Value", "New Value", "Source / Reason"]
for i, h in enumerate(headers, 1):
    c = ws_log.cell(row=4, column=i, value=h)
    c.font = Font(bold=True, name="Calibri", size=9)
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.font = Font(bold=True, name="Calibri", size=9, color="FFFFFF")

corrections = [
    ("Assumptions!C10", "$78,000", "$100,688", "BTC anchor: CIFR Q1'26 realized ($34,838K rev ÷ 346 BTC). Prior $78k was stale."),
    ("CIFR!C7",         "0 MW",    "493 MW",   "Signed-not-energized: Barber Lake 207MW (Fluidstack) + Black Pearl 216MW (Amazon) + Stingray 70MW. Q1'26 10-Q project disclosures."),
    ("CIFR!C10",        "700 MW",  "207 MW",   "Pipeline: signed leases moved to C7; ~207 MW remaining unsigned/dev capacity."),
    ("CIFR!C12",        "$70,820K","$72,000K", "Mining cost: Q1'26 10-Q CoR $18,000K/q × 4 = $72,000K/yr."),
    ("CIFR!C13",        "792 BTC", "1,116 BTC","BTC treasury: Q1'26 10-Q 'digital assets held 1,116 BTC at FV $111.5M'. Prior 792 was stale."),
]

still_est = [
    ("WULF!C15",  "$80,000K SBC",  "UNVERIFIED","FY25 10-K SG&A cites ~$20M vs Q1'26 cash-flow $101M. EDGAR access blocked — VERIFY."),
    ("CORZ!C14",  "$150,000K SG&A","ESTIMATE",  "SG&A not confirmed from SEC filing. Mark ESTIMATE until verified."),
    ("CLSK!C14",  "$120,000K SG&A","ESTIMATE",  "SG&A not confirmed from SEC filing."),
    ("CLSK!C15",  "$70,000K SBC",  "ESTIMATE",  "SBC not confirmed from SEC filing."),
    ("HUT!C7",    "0 MW",          "UNVERIFIED","HUT signed AI DC leases ($16.8B TCV) → MW equivalent not confirmed from SEC filing. Needs 10-Q review."),
]

for i, row in enumerate(corrections, 5):
    for j, val in enumerate(row, 1):
        c = ws_log.cell(row=i, column=j, value=val)
        c.font = Font(name="Calibri", size=9)
        c.fill = PatternFill("solid", fgColor="C6EFCE")
        c.alignment = Alignment(wrap_text=True)

ws_log.cell(row=11, column=1).value = "STILL NEEDS VERIFICATION (EDGAR blocked — cannot pull directly):"
ws_log.cell(row=11, column=1).font = Font(bold=True, name="Calibri", size=9, color="C00000")

for i, row in enumerate(still_est, 12):
    for j, val in enumerate(row, 1):
        c = ws_log.cell(row=i, column=j, value=val)
        c.font = Font(name="Calibri", size=9)
        c.fill = PatternFill("solid", fgColor="FCE4D6")
        c.alignment = Alignment(wrap_text=True)

ws_log.row_dimensions[4].height = 16
for r in range(5, 18):
    ws_log.row_dimensions[r].height = 30

wb.save(DST)
print(f"Saved: {DST}")

# Print impact summary using openpyxl data_only to read formulas vs values
wb2 = openpyxl.load_workbook(DST, data_only=True)
print("\n=== CORRECTION IMPACT ON KEY OUTPUTS ===")
print("(formula cells show None in data_only mode if not pre-calculated)")

for co in ["CIFR", "WULF", "CORZ", "HUT", "CLSK", "RIOT"]:
    ws = wb2[co]
    intrinsic_ev = ws["C42"].value
    intrinsic_eq = ws["C43"].value
    per_share    = ws["C44"].value
    upside       = ws["C50"].value
    print(f"\n  {co}:")
    print(f"    Intrinsic EV  (C42): {intrinsic_ev}")
    print(f"    Intrinsic Eq  (C43): {intrinsic_eq}")
    print(f"    Per share     (C44): {per_share}")
    print(f"    Upside vs mkt (C50): {upside}")
