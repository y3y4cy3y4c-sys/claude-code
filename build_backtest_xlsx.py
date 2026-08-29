#!/usr/bin/env python3
"""Build 5s30s_Backtest_TenorComparison.xlsx from the backtest dataset.

Sheets:
  1. Dashboard          - headline findings (formulas linked to Cycle Backtest)
  2. Cycle Backtest     - per-cycle results, all computed by formula
  3. Yield Data         - checkpoint yields (blue inputs) + spread/delta formulas
  4. Tenor Comparison   - 2s30s/5s30s/10s30s on the Aug-27-26 curve (duration approx)
Conventions: blue = hardcoded input, black = formula, green = cross-sheet link.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from backtest_5s30s_easing_cycles import EMBEDDED_DATA

ARIAL = "Arial"
F_TITLE = Font(name=ARIAL, size=13, bold=True)
F_H = Font(name=ARIAL, size=9, bold=True)
F_B = Font(name=ARIAL, size=10)
F_IN = Font(name=ARIAL, size=10, color="0000FF")          # blue input
F_LINK = Font(name=ARIAL, size=10, color="008000")        # green cross-sheet
F_NOTE = Font(name=ARIAL, size=9, italic=True, color="666666")
FILL_H = PatternFill("solid", fgColor="1F3B34")
F_HW = Font(name=ARIAL, size=9, bold=True, color="FFFFFF")
FILL_KEY = PatternFill("solid", fgColor="FFFF00")
THIN = Border(bottom=Side(style="thin", color="C8C8C8"))

BP = '+0;(0);"-"'
BP_PLAIN = '0;(0);"-"'
PCT2 = "0.00"

wb = Workbook()

# ---------------------------------------------------------------- Yield Data
ws = wb.create_sheet("Yield Data")
ws["A1"] = "Checkpoint yields around Fed easing cycles (CMT, %)"
ws["A1"].font = F_TITLE
ws["A2"] = ("BLUE = input data. Quarterly month-end approximations reconstructed from the historical record "
            "(primary source is FRED H.15 daily CMTs; this sandbox had no network access to FRED, so values carry "
            "the per-cycle tolerance flagged on 'Cycle Backtest'). Peaks cross-checked against published anchors: "
            "post-GFC 5s30s peak ~+280-300bp; Jun-2025 +101bp 'steepest since 2021'. The +23m row of the 2024-25 "
            "cycle is official H.15 Aug-27-2026.")
ws["A2"].font = F_NOTE
hdrs = ["Cycle", "Offset (m)", "2y", "5y", "10y", "30y",
        "2s30s (bp)", "5s30s (bp)", "10s30s (bp)",
        "d2s30s vs T0", "d5s30s vs T0", "d10s30s vs T0"]
for j, h in enumerate(hdrs, 1):
    c = ws.cell(row=3, column=j, value=h)
    c.font = F_HW
    c.fill = FILL_H
r = 4
meta = {}
for name, cyc in EMBEDDED_DATA.items():
    start = r
    offs = sorted(cyc["path"])
    t0row = None
    rows_by_off = {}
    for off in offs:
        y2, y5, y10, y30 = cyc["path"][off]
        ws.cell(row=r, column=1, value=name).font = F_B
        ws.cell(row=r, column=2, value=off).font = F_B
        for j, v in enumerate((y2, y5, y10, y30), 3):
            c = ws.cell(row=r, column=j, value=v)
            c.font = F_IN
            c.number_format = PCT2
        if off == 0:
            t0row = r
        rows_by_off[off] = r
        r += 1
    for off in offs:
        rr = rows_by_off[off]
        ws.cell(row=rr, column=7, value=f"=(F{rr}-C{rr})*100")
        ws.cell(row=rr, column=8, value=f"=(F{rr}-D{rr})*100")
        ws.cell(row=rr, column=9, value=f"=(F{rr}-E{rr})*100")
        ws.cell(row=rr, column=10, value=f"=G{rr}-G${t0row}")
        ws.cell(row=rr, column=11, value=f"=H{rr}-H${t0row}")
        ws.cell(row=rr, column=12, value=f"=I{rr}-I${t0row}")
        for j in range(7, 13):
            ws.cell(row=rr, column=j).font = F_B
            ws.cell(row=rr, column=j).number_format = BP_PLAIN if j < 10 else BP
    for j in range(1, 13):
        ws.cell(row=r - 1, column=j).border = THIN
    meta[name] = dict(start=start, end=r - 1, t0=t0row, rows=rows_by_off)
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 10
for col in "CDEF":
    ws.column_dimensions[col].width = 8
for col in ("G", "H", "I", "J", "K", "L"):
    ws.column_dimensions[col].width = 12
ws.freeze_panes = "A4"

# ------------------------------------------------------------- Cycle Backtest
bt = wb.create_sheet("Cycle Backtest")
bt["A1"] = "5s30s steepener across Fed easing cycles — entry at first cut (T0), bp of curve"
bt["A1"].font = F_TITLE
bt["A2"] = "All results computed by formula from 'Yield Data'. Quarterly checkpoints understate intra-month peaks slightly."
bt["A2"].font = F_NOTE
cols = ["Cycle", "First cut", "Cycle kind", "Data quality", "Funds start %", "Funds trough %",
        "Cuts (bp)", "5s30s @T0", "T-6m to T0 gain", "D +6m", "D +12m", "D +24m",
        "Peak D", "@ month", "Worst D", "Peak D 2s30s", "Peak D 10s30s"]
for j, h in enumerate(cols, 1):
    c = bt.cell(row=4, column=j, value=h)
    c.font = F_HW
    c.fill = FILL_H
FMT = {5: PCT2, 6: PCT2, 7: BP_PLAIN, 8: BP_PLAIN, 9: BP, 10: BP, 11: BP,
       12: BP, 13: BP, 14: BP_PLAIN, 15: BP, 16: BP, 17: BP}
row = 5
first_data_row = row
deep_rows = []
YD = "'Yield Data'"
for name, cyc in EMBEDDED_DATA.items():
    m = meta[name]
    pos0, endr = m["t0"], m["end"]
    bt.cell(row=row, column=1, value=name).font = F_B
    bt.cell(row=row, column=2, value=cyc["first_cut"]).font = F_IN
    bt.cell(row=row, column=3, value=cyc["kind"]).font = F_IN
    bt.cell(row=row, column=4, value=cyc["quality"]).font = F_IN
    bt.cell(row=row, column=5, value=cyc["funds_start"]).font = F_IN
    bt.cell(row=row, column=6, value=cyc["funds_trough"]).font = F_IN
    bt.cell(row=row, column=7, value=f"=(E{row}-F{row})*100")
    bt.cell(row=row, column=8, value=f"={YD}!H{pos0}")
    bt.cell(row=row, column=9, value=f"={YD}!H{pos0}-{YD}!H{m['rows'][-6]}")
    for j, off in ((10, 6), (11, 12), (12, 24)):
        if off in m["rows"]:
            bt.cell(row=row, column=j, value=f"={YD}!K{m['rows'][off]}")
        else:
            bt.cell(row=row, column=j, value="n/a").font = F_NOTE
    bt.cell(row=row, column=13, value=f"=MAX({YD}!K{pos0}:K{endr})")
    bt.cell(row=row, column=14,
            value=f"=INDEX({YD}!B{pos0}:B{endr},MATCH(M{row},{YD}!K{pos0}:K{endr},0))")
    bt.cell(row=row, column=15, value=f"=MIN({YD}!K{pos0}:K{endr})")
    bt.cell(row=row, column=16, value=f"=MAX({YD}!J{pos0}:J{endr})")
    bt.cell(row=row, column=17, value=f"=MAX({YD}!L{pos0}:L{endr})")
    for j in range(7, 18):
        cc = bt.cell(row=row, column=j)
        if cc.value != "n/a":
            cc.font = F_LINK if j in (8, 9, 10, 11, 12) else F_B
            cc.number_format = FMT[j]
    for j in (5, 6):
        bt.cell(row=row, column=j).number_format = FMT[j]
    if "deep" in cyc["kind"] or "COVID" in cyc["kind"]:
        deep_rows.append(row)
    row += 1
last_data_row = row - 1
row += 1
bt.cell(row=row, column=1, value="SUMMARY").font = F_H
srow = row + 1
summaries = [
    ("Median D +12m (all cycles)", f"=MEDIAN(K{first_data_row}:K{last_data_row})", BP),
    ("Hit rate: D +12m > 0",
     f'=COUNTIF(K{first_data_row}:K{last_data_row},">0")&" of "&COUNT(K{first_data_row}:K{last_data_row})', None),
    ("Hit rate: D +12m >= +30bp",
     f'=COUNTIF(K{first_data_row}:K{last_data_row},">=30")&" of "&COUNT(K{first_data_row}:K{last_data_row})', None),
    ("Worst D +12m (1998 re-hike)", f"=MIN(K{first_data_row}:K{last_data_row})", BP),
    ("Median peak D, deep cycles", "=MEDIAN(" + ",".join(f"M{r0}" for r0 in deep_rows) + ")", BP),
    ("Median months to peak (all)", f"=MEDIAN(N{first_data_row}:N{last_data_row})", BP_PLAIN),
    ("Median T-6m to T0 gain (pre-positioning)", f"=MEDIAN(I{first_data_row}:I{last_data_row})", BP),
    ("2s30s peak beta vs 5s30s (deep median)", "=MEDIAN(" + ",".join(f"P{r0}/M{r0}" for r0 in deep_rows) + ")", '0.00"x"'),
    ("10s30s peak beta vs 5s30s (deep median)", "=MEDIAN(" + ",".join(f"Q{r0}/M{r0}" for r0 in deep_rows) + ")", '0.00"x"'),
]
for i, (label, formula, fmt) in enumerate(summaries):
    bt.cell(row=srow + i, column=1, value=label).font = F_B
    c = bt.cell(row=srow + i, column=2, value=formula)
    c.font = F_B
    if fmt:
        c.number_format = fmt
bt.column_dimensions["A"].width = 30
bt.column_dimensions["B"].width = 12
bt.column_dimensions["C"].width = 40
bt.column_dimensions["D"].width = 28
for col in "EFGHIJKLMNOPQ":
    bt.column_dimensions[col].width = 11
bt.freeze_panes = "A5"

# ------------------------------------------------------------ Tenor Comparison
tc = wb.create_sheet("Tenor Comparison")
tc["A1"] = "2s30s vs 5s30s vs 10s30s — actual curve, Aug 27-28 2026"
tc["A1"].font = F_TITLE
tc["A2"] = ("Curve: official H.15 Aug-27-2026 where available; * = interpolated estimate (blue = editable input). "
            "Carry+roll here uses the duration approximation; carry_roll_tenor_comparison.py reprices aged par bonds "
            "exactly (differences <2bp/yr). At the 2y, static carry is mostly the priced hike premium — realized only "
            "if the Fed does not hike: there, carry IS the trade, not a bonus.")
tc["A2"].font = F_NOTE
tc["A2"].alignment = Alignment(wrap_text=True, vertical="top")
tc.merge_cells("A2:I3")
tc["A5"] = "Inputs"
tc["A5"].font = F_H
curve = [("1y *", 4.08), ("2y", 4.24), ("3y *", 4.29), ("5y", 4.38), ("7y *", 4.51),
         ("10y", 4.67), ("20y", 5.18), ("30y", 5.19)]
tc.cell(row=6, column=1, value="Tenor").font = F_HW
tc.cell(row=6, column=1).fill = FILL_H
tc.cell(row=6, column=2, value="Yield %").font = F_HW
tc.cell(row=6, column=2).fill = FILL_H
crow = {}
for i, (t, y) in enumerate(curve):
    rr = 7 + i
    tc.cell(row=rr, column=1, value=t).font = F_B
    c = tc.cell(row=rr, column=2, value=y)
    c.font = F_IN
    c.number_format = PCT2
    crow[t.split()[0]] = rr
tc.cell(row=15, column=1, value="Repo / funding %").font = F_B
c = tc.cell(row=15, column=2, value=3.65)
c.font = F_IN
c.number_format = PCT2
c.fill = FILL_KEY
tc.cell(row=16, column=1, value="DV01 per leg ($/bp)").font = F_B
c = tc.cell(row=16, column=2, value=20000)
c.font = F_IN
c.number_format = '$#,##0'
c.fill = FILL_KEY
tc.cell(row=17, column=1, value="Mod. duration, par bond: 2y / 5y / 10y / 30y").font = F_B
for j, d in enumerate((1.90, 4.45, 8.09, 15.17), 3):
    c = tc.cell(row=17, column=j, value=d)
    c.font = F_IN
    c.number_format = "0.00"
tc.cell(row=17, column=7, value="semiannual par-bond pricing at input yields (see .py)").font = F_NOTE
tc.cell(row=18, column=1, value="Aged mod. duration: 1y / 4y / 9y / 29y").font = F_B
for j, d in enumerate((0.97, 3.63, 7.36, 14.90), 3):
    c = tc.cell(row=18, column=j, value=d)
    c.font = F_IN
    c.number_format = "0.00"
tc.cell(row=19, column=1, value="Aged yield (interp): 1y / 4y / 9y / 29y").font = F_B
aged_fs = [f"=B{crow['1y']}",
           f"=(B{crow['3y']}+B{crow['5y']})/2",
           f"=B{crow['7y']}+(B{crow['10y']}-B{crow['7y']})*2/3",
           f"=B{crow['20y']}+(B{crow['30y']}-B{crow['20y']})*0.9"]
for j, fml in enumerate(aged_fs, 3):
    c = tc.cell(row=19, column=j, value=fml)
    c.font = F_B
    c.number_format = PCT2

hdr2 = ["Pair", "Front yield %", "Front face $mm", "30y face $mm", "Gross $mm",
        "Running carry $k/yr", "Roll $k/yr", "Carry+roll bp/yr", "Deep-cycle peak beta"]
for j, h in enumerate(hdr2, 1):
    c = tc.cell(row=21, column=j, value=h)
    c.font = F_HW
    c.fill = FILL_H
# per-pair cells: front tenor row, duration cell (row17), aged duration cell (row18), aged yield cell (row19)
pairs = [("2s30s", "2y", "C"), ("5s30s", "5y", "D"), ("10s30s", "10y", "E")]
beta_link = {"2s30s": f"='Cycle Backtest'!B{srow + 7}", "10s30s": f"='Cycle Backtest'!B{srow + 8}"}
for i, (pname, ten, col) in enumerate(pairs):
    rr = 22 + i
    yc = f"B{crow[ten]}"
    y30 = f"B{crow['30y']}"
    tc.cell(row=rr, column=1, value=pname).font = F_B
    c = tc.cell(row=rr, column=2, value=f"={yc}")
    c.number_format = PCT2
    tc.cell(row=rr, column=3, value=f"=$B$16/({col}17*100)").number_format = "0.0"
    tc.cell(row=rr, column=4, value="=$B$16/($F$17*100)").number_format = "0.0"
    tc.cell(row=rr, column=5, value=f"=C{rr}+D{rr}").number_format = "0.0"
    tc.cell(row=rr, column=6,
            value=f"=(({yc}-$B$15)*C{rr}-({y30}-$B$15)*D{rr})*10").number_format = '#,##0'
    tc.cell(row=rr, column=7,
            value=f"=(({yc}-{col}19)*{col}18*C{rr}-({y30}-F19)*F18*D{rr})*10").number_format = '#,##0'
    tc.cell(row=rr, column=8, value=f"=(F{rr}+G{rr})*1000/$B$16").number_format = "+0.0;(0.0)"
    if pname in beta_link:
        c = tc.cell(row=rr, column=9, value=beta_link[pname])
        c.font = F_LINK
    else:
        c = tc.cell(row=rr, column=9, value=1.00)
        c.font = F_IN
    c.number_format = '0.00"x"'
    for j in (2, 3, 4, 5, 6, 7, 8):
        tc.cell(row=rr, column=j).font = F_B
for col, w in (("A", 13), ("B", 12), ("C", 13), ("D", 12), ("E", 11), ("F", 17), ("G", 12), ("H", 16), ("I", 18)):
    tc.column_dimensions[col].width = w
tc.cell(row=26, column=1,
        value=("Reading: 5s30s delivers ~62% of 2s30s's historical cyclical payoff on ~half the balance sheet, with "
               "carry that is positive without being purely the hike-premium bet; 10s30s keeps only ~1/3 of the payoff. "
               "Futures legs (FV vs Ultra Bond) should be sized dynamically off current CTD DV01s, not a memorized ratio.")).font = F_NOTE
tc.cell(row=26, column=1).alignment = Alignment(wrap_text=True, vertical="top")
tc.merge_cells("A26:I28")

# ---------------------------------------------------------------- Dashboard
db = wb.create_sheet("Dashboard", 0)
db["A1"] = "5s30s Steepener — Easing-Cycle Backtest & Tenor Validation"
db["A1"].font = Font(name=ARIAL, size=15, bold=True)
db["A2"] = "Companion quant workbook to the Q4 trade memo (macro-trade-framework branch)"
db["A2"].font = F_NOTE
db["A3"] = "Last updated: 2026-08-29 (curve as of H.15 Aug-27/28-2026)"
db["A3"].font = F_NOTE
db["A5"] = "HEADLINE FINDINGS (live links to 'Cycle Backtest')"
db["A5"].font = F_H
find = [
    ("Cycles tested (1984-2025)", f"=COUNTA('Cycle Backtest'!A{first_data_row}:A{last_data_row})", "0"),
    ("Hit rate, D+12m > 0", f"='Cycle Backtest'!B{srow + 1}", None),
    ("Median D+12m, all cycles (bp)", f"='Cycle Backtest'!B{srow}", BP),
    ("Median peak D, deep cycles (bp)", f"='Cycle Backtest'!B{srow + 4}", BP),
    ("Median months to peak", f"='Cycle Backtest'!B{srow + 5}", BP_PLAIN),
    ("Worst 12m outcome (1998: cuts then re-hike)", f"='Cycle Backtest'!B{srow + 3}", BP),
    ("Median pre-positioning gain (T-6m to T0)", f"='Cycle Backtest'!B{srow + 6}", BP),
    ("2s30s / 10s30s peak beta vs 5s30s",
     f"=TEXT('Cycle Backtest'!B{srow + 7},\"0.00\")&\"x / \"&TEXT('Cycle Backtest'!B{srow + 8},\"0.00\")&\"x\"", None),
]
rr = 6
for label, formula, fmt in find:
    db.cell(row=rr, column=1, value=label).font = F_B
    c = db.cell(row=rr, column=2, value=formula)
    c.font = F_LINK
    if fmt:
        c.number_format = fmt
    rr += 1
db["A15"] = "WHAT THE BACKTEST SAYS ABOUT THE TRADE"
db["A15"].font = F_H
lessons = [
    "1. Steepeners are cycle trades: every deep easing cycle (1989-92, 2001-03, 2007-10, and the COVID extension) "
    "delivered +180 to +230bp peaks vs T0 — but the median peak arrives ~18 months after the first cut. Positive carry matters.",
    "2. The failure mode is the shallow/insurance cycle: 1995 went nowhere; 1998 lost 50bp within 12m when the Fed re-hiked. "
    "That is exactly the scenario the memo's stop (+45-50bp level) and Board-vote falsifier are built around.",
    "3. Pre-positioning before the first cut historically paid (median +33bp from T-6m into T0) — but today's entry is 23 months "
    "into a PAUSED cycle with hike pricing: the honest analogs split between 1990 (pause, then recession, then mega-steepening) "
    "and 1998-99 (re-hike, then flattening). The labor market is the differentiator — per the memo's Q1.",
    "4. Target calibration: initial target +110-120 means regaining the 2025 steepness (peak ~+110) — requires only that cuts "
    "resume, not a recession. The +150 stretch sits BELOW every deep-cycle peak LEVEL (+170 to +295) — conservative if the "
    "cycle turns out deep.",
    "5. Tenor choice: 2s30s captured ~1.6x of 5s30s's peak but needs ~2x the balance sheet and its carry IS the Fed bet; "
    "10s30s kept only ~0.33x. 5s30s: ~62% of the payoff, ~half the gross, positive carry that isn't purely the hike premium.",
]
rr = 16
for t in lessons:
    db.cell(row=rr, column=1, value=t).font = F_B
    db.cell(row=rr, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    db.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)
    db.row_dimensions[rr].height = 42
    rr += 1
db["A22"] = "REVISED TRADE PARAMETERS (post-validation)"
db["A22"].font = F_H
params = [("Entry (5s30s)", "+75 to +80bp (H.15 Aug-27: +81; post-Warsh flattening)"),
          ("Initial target", "+110 to +120bp"),
          ("Stretch target", "+150bp (deep-cycle scenario only)"),
          ("Stop", "+45 to +50bp"),
          ("Sizing", "DV01-neutral, $20k/bp per leg; futures legs sized dynamically off current CTD DV01s"),
          ("Carry + roll", "~ +10bp/yr (aged-bond repricing, curve unchanged; assumption-sensitive, sign robust)")]
rr = 23
for k, v in params:
    db.cell(row=rr, column=1, value=k).font = F_B
    db.cell(row=rr, column=2, value=v).font = F_IN
    rr += 1
db["A30"] = ("Method & caveats: quarterly checkpoints understate intra-month extremes; 1984-86 data +/-25-30bp; 30y after "
             "Feb-2002 is a long-bond proxy (issuance gap). Scenario probabilities in the memo are subjective priors, not "
             "outputs of this backtest. Re-run backtest_5s30s_easing_cycles.py against daily FRED DGS2/DGS5/DGS10/DGS30 "
             "CSVs for exact levels when network access allows.")
db["A30"].font = F_NOTE
db["A30"].alignment = Alignment(wrap_text=True, vertical="top")
db.merge_cells("A30:F32")
db.column_dimensions["A"].width = 44
db.column_dimensions["B"].width = 62

wb.remove(wb["Sheet"])
wb.save("5s30s_Backtest_TenorComparison.xlsx")
print("saved 5s30s_Backtest_TenorComparison.xlsx")
