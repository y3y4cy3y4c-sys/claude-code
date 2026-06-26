"""
Builds npv_model.xlsx — a single-sheet Excel NPV model for HPC data center projects.
Supports Colocation (CIFR/WULF) and AI Cloud (IREN) project types.
Upload directly to Google Sheets.
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "NPV Model"

# ─────────────────────────────────────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────────────────────────────────────

# Colors
C_NAVY      = "1F3864"
C_BLUE      = "2E5EAA"
C_LIGHTBLUE = "BDD7EE"
C_MIDBLUE   = "9DC3E6"
C_YELLOW    = "FFD966"
C_GREEN     = "E2EFDA"
C_RED       = "FCE4D6"
C_GREY      = "F2F2F2"
C_WHITE     = "FFFFFF"
C_ORANGE    = "F4B942"
C_DARKGREY  = "595959"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_bottom(color="AAAAAA"):
    s = Side(style="thin", color=color)
    return Border(bottom=s)

def border_box():
    s = Side(style="thin", color="AAAAAA")
    return Border(top=s, bottom=s, left=s, right=s)

def border_thick_bottom():
    s = Side(style="medium", color=C_NAVY)
    return Border(bottom=s)

FMT_MM    = '#,##0.0"mm"'
FMT_PCT   = '0.0%'
FMT_INT   = '#,##0'
FMT_MMPCT = '#,##0.0"mm"'

# ─────────────────────────────────────────────────────────────────────────────
# COLUMN WIDTHS
# ─────────────────────────────────────────────────────────────────────────────

col_widths = {
    "A": 3,    # spacer
    "B": 28,   # label
    "C": 3,    # spacer
    "D": 16,   # project 1
    "E": 16,   # project 2
    "F": 16,   # project 3
    "G": 16,   # project 4
    "H": 16,   # project 5
    "I": 16,   # project 6
    "J": 3,    # spacer
    "K": 18,   # total / notes
}

for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# ─────────────────────────────────────────────────────────────────────────────
# HELPER: write a styled cell
# ─────────────────────────────────────────────────────────────────────────────

def w(row, col, value, bg=None, fg="000000", bold=False, size=10,
      fmt=None, h_align="left", italic=False, wrap=False, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=size, italic=italic)
    c.alignment = align(h=h_align, wrap=wrap)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = border
    return c

def merge_w(row, col1, col2, value, bg=None, fg="000000", bold=False,
            size=10, h_align="center", wrap=False, italic=False):
    ws.merge_cells(start_row=row, start_column=col1,
                   end_row=row,   end_column=col2)
    c = ws.cell(row=row, column=col1, value=value)
    if bg:
        c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=size, italic=italic)
    c.alignment = align(h=h_align, wrap=wrap)
    return c

def fill_row(row, col1, col2, bg):
    for col in range(col1, col2 + 1):
        ws.cell(row=row, column=col).fill = fill(bg)

PROJECT_COLS = [4, 5, 6, 7, 8, 9]   # D..I
DATA_COLS    = list(range(2, 12))    # B..K

# ─────────────────────────────────────────────────────────────────────────────
# ROW LAYOUT — build top-to-bottom
# ─────────────────────────────────────────────────────────────────────────────

R = 1  # current row tracker

# ── TITLE BAR ────────────────────────────────────────────────────────────────
ws.row_dimensions[R].height = 28
merge_w(R, 2, 11, "HPC DATA CENTER  —  NPV MODEL",
        bg=C_NAVY, fg=C_WHITE, bold=True, size=14)
fill_row(R, 1, 11, C_NAVY)
R += 1

ws.row_dimensions[R].height = 14
merge_w(R, 2, 11,
        "CIFR (Cipher Digital) — 3 Signed Projects + 3 Pipeline Sites  |  Base Year: 2026  |  All figures in $mm USD",
        bg=C_BLUE, fg=C_WHITE, size=9, italic=True)
fill_row(R, 1, 11, C_BLUE)
R += 1

R += 1  # blank

# ── SECTION: PROJECT HEADERS ──────────────────────────────────────────────────
ws.row_dimensions[R].height = 10
R += 1

ws.row_dimensions[R].height = 20
merge_w(R, 2, 3, "PROJECT", bg=C_NAVY, fg=C_WHITE, bold=True, size=9, h_align="left")
for i, col in enumerate(PROJECT_COLS):
    label = f"Project {i+1}"
    w(R, col, label, bg=C_NAVY, fg=C_YELLOW, bold=True, size=9, h_align="center")
w(R, 10, "NOTES", bg=C_NAVY, fg=C_WHITE, bold=True, size=9, h_align="left")
R += 1

ROW_TYPE    = R
ws.row_dimensions[R].height = 18
w(R, 2, "Project Type", bg=C_LIGHTBLUE, bold=True, size=9)
for col in PROJECT_COLS:
    c = ws.cell(row=R, column=col)
    c.fill = fill(C_YELLOW)
    c.font = font(bold=True, size=9)
    c.alignment = align(h="center")
    # Dropdown hint via comment — actual data validation needs openpyxl.worksheet.datavalidation
w(R, 10, 'Type: "Colo" or "Cloud"', bg=C_GREY, fg=C_DARKGREY, size=8, italic=True)
R += 1

ROW_COMPANY = R
ws.row_dimensions[R].height = 16
w(R, 2, "Company", bg=C_LIGHTBLUE, bold=True, size=9)
for col in PROJECT_COLS:
    ws.cell(row=R, column=col).fill = fill(C_YELLOW)
    ws.cell(row=R, column=col).font = font(bold=True, size=9)
    ws.cell(row=R, column=col).alignment = align(h="center")
R += 1

ROW_PROJECT_NAME = R
ws.row_dimensions[R].height = 16
w(R, 2, "Project Name", bg=C_LIGHTBLUE, bold=True, size=9)
for col in PROJECT_COLS:
    ws.cell(row=R, column=col).fill = fill(C_YELLOW)
    ws.cell(row=R, column=col).font = font(bold=True, size=9)
    ws.cell(row=R, column=col).alignment = align(h="center")
R += 1

ROW_STATUS = R
ws.row_dimensions[R].height = 16
w(R, 2, "Status", bg=C_LIGHTBLUE, size=9)
for col in PROJECT_COLS:
    ws.cell(row=R, column=col).fill = fill(C_LIGHTBLUE)
    ws.cell(row=R, column=col).font = font(size=8, italic=True, color=C_DARKGREY)
    ws.cell(row=R, column=col).alignment = align(h="center", wrap=True)
R += 1

R += 1  # blank

# ── SECTION A: REVENUE INPUTS ─────────────────────────────────────────────────
def section_header(row, label, note=""):
    ws.row_dimensions[row].height = 18
    merge_w(row, 2, 3, f"  ▸  {label}", bg=C_BLUE, fg=C_WHITE, bold=True, size=9, h_align="left")
    for col in PROJECT_COLS:
        ws.cell(row=row, column=col).fill = fill(C_BLUE)
    w(row, 10, note, bg=C_BLUE, fg=C_WHITE, size=8, italic=True)

def input_row(row, label, fmt=None, note="", height=15, bg=C_WHITE):
    ws.row_dimensions[row].height = height
    w(row, 2, label, bg=bg, size=9)
    for col in PROJECT_COLS:
        c = ws.cell(row=row, column=col)
        c.fill = fill(C_YELLOW)
        c.font = font(size=9)
        c.alignment = align(h="center")
        if fmt:
            c.number_format = fmt
    w(row, 10, note, bg=C_GREY, fg=C_DARKGREY, size=8, italic=True)
    return row

def calc_row(row, label, fmt=None, note="", height=15, bg=C_GREEN):
    ws.row_dimensions[row].height = height
    w(row, 2, f"  {label}", bg=bg, size=9)
    for col in PROJECT_COLS:
        c = ws.cell(row=row, column=col)
        c.fill = fill(C_GREEN)
        c.font = font(size=9, color="1A6630")
        c.alignment = align(h="center")
        if fmt:
            c.number_format = fmt
    w(row, 10, note, bg=C_GREY, fg=C_DARKGREY, size=8, italic=True)
    return row

def output_row(row, label, fmt=None, note="", bold=False, height=16, bg=C_GREY):
    ws.row_dimensions[row].height = height
    w(row, 2, f"  {label}", bg=bg, size=9, bold=bold)
    for col in PROJECT_COLS:
        c = ws.cell(row=row, column=col)
        c.fill = fill(bg)
        c.font = font(size=9, bold=bold)
        c.alignment = align(h="center")
        if fmt:
            c.number_format = fmt
    w(row, 10, note, bg=C_GREY, fg=C_DARKGREY, size=8, italic=True)
    return row

section_header(R, "A.  REVENUE INPUTS", "Fill yellow cells. Green cells = auto-calculated.")
R += 1

ROW_MW          = input_row(R, "Critical IT Load (MW)", FMT_INT, "Net MW for colocation; or leave blank for Cloud"); R += 1
ROW_REV_PW      = input_row(R, "Revenue per Watt per Year ($/W/yr)", '"$"#,##0.000', "Colo: ~$1.50–1.65/W/yr  |  Cloud: leave blank"); R += 1
ROW_ARR         = input_row(R, "Contracted ARR ($mm/yr)", FMT_MM, "Cloud projects: enter annual contracted revenue here"); R += 1
ROW_ANNUAL_REV  = calc_row(R,  "Annual Revenue ($mm/yr)", FMT_MM, "=MW × $/W/yr  OR  =ARR (whichever applies)"); R += 1
ROW_UTIL        = input_row(R, "Utilization Rate (%)", FMT_PCT, "Cloud only. Colo: enter 100%"); R += 1
ROW_REALIZED    = calc_row(R,  "Realized Revenue ($mm/yr)", FMT_MM, "= Annual Revenue × Utilization"); R += 1
ROW_NOI_MARGIN  = input_row(R, "NOI / EBITDA Margin (%)", FMT_PCT, "Colo NNN: ~85%  |  Cloud: compute from inputs below"); R += 1
ROW_ANNUAL_NOI  = calc_row(R,  "Annual NOI / EBITDA ($mm/yr)", FMT_MM, "= Realized Revenue × Margin"); R += 1

R += 1  # blank

# ── SECTION B: COST INPUTS (Cloud) ───────────────────────────────────────────
section_header(R, "B.  CLOUD COST INPUTS  (IREN-style only; leave blank for Colo)")
R += 1

ROW_POWER       = input_row(R, "  Power Cost ($mm/yr)", FMT_MM, "Cloud: est. from MW × $/kWh"); R += 1
ROW_DC_OPEX     = input_row(R, "  DC Opex ($mm/yr)", FMT_MM); R += 1
ROW_SW_SUPPORT  = input_row(R, "  Software & Support ($mm/yr)", FMT_MM); R += 1
ROW_INTEREST    = input_row(R, "  Interest on Debt ($mm/yr)", FMT_MM, "= Debt × Interest Rate"); R += 1
ROW_CLOUD_EBITDA= calc_row(R,  "  Cloud EBITDA ($mm/yr)", FMT_MM, "= Revenue - Power - DC Opex - SW"); R += 1
ROW_CLOUD_FCF   = calc_row(R,  "  Cloud FCF post-interest ($mm/yr)", FMT_MM, "= EBITDA - Interest"); R += 1

R += 1  # blank

# ── SECTION C: TIMING & DISCOUNT ─────────────────────────────────────────────
section_header(R, "C.  TIMING & DISCOUNT RATE")
R += 1

ROW_COD         = input_row(R, "COD Year", FMT_INT, "Year cash flow begins"); R += 1
ROW_YRS_TO_COD  = calc_row(R,  "Years to COD", FMT_INT, "= COD Year - 2026"); R += 1
ROW_TERM        = input_row(R, "Lease / Contract Term (years)", FMT_INT); R += 1
ROW_DISC        = input_row(R, "Discount Rate (WACC)", FMT_PCT, "Colo: ~10%  |  Cloud: ~12%"); R += 1
ROW_ANN_FACTOR  = calc_row(R,  "Annuity Factor", '"×"#,##0.000', "= PV annuity at discount rate for term"); R += 1
ROW_DISC_FACTOR = calc_row(R,  "Discount Factor to Today", '"×"#,##0.000', "= 1/(1+r)^years_to_COD"); R += 1

R += 1  # blank

# ── SECTION D: CAPEX & FINANCING ─────────────────────────────────────────────
section_header(R, "D.  CAPEX & FINANCING")
R += 1

ROW_CAPEX       = input_row(R, "Remaining Construction / GPU Capex ($mm)", FMT_MM, "Colo: build cost  |  Cloud: GPU purchase cost"); R += 1
ROW_LTC         = input_row(R, "Loan-to-Cost Ratio (%)", FMT_PCT, "Typically 85% for colo projects"); R += 1
ROW_DEBT        = input_row(R, "Project Debt ($mm)", FMT_MM, "Can auto-calc: = Capex × LTC, or enter directly"); R += 1
ROW_EQUITY      = calc_row(R,  "Equity Invested ($mm)", FMT_MM, "= Capex × (1 - LTC)"); R += 1

R += 1  # blank

# ── SECTION E: GPU REPLACEMENT (Cloud only) ───────────────────────────────────
section_header(R, "E.  GPU REPLACEMENT CAPEX  (Cloud only; leave blank for Colo)")
R += 1

ROW_GPU_LIFE    = input_row(R, "  GPU Useful Life (years)", FMT_INT, "Typically 4 years for Blackwell-class"); R += 1
ROW_REPL_PCT    = input_row(R, "  Replacement Capex (% of initial)", FMT_PCT, "~80% — next-gen GPUs cost similar"); R += 1
ROW_REPL_COST   = calc_row(R,  "  Replacement Capex Amount ($mm)", FMT_MM, "= Capex × Replacement %"); R += 1
ROW_REPL_YR     = calc_row(R,  "  Replacement Year", FMT_INT, "= COD Year + GPU Life"); R += 1
ROW_PV_REPL     = calc_row(R,  "  PV of Replacement Capex ($mm)", FMT_MM, "Discounted back to today"); R += 1

R += 1  # blank

# ── SECTION F: EXTENSION / RENEWAL ───────────────────────────────────────────
section_header(R, "F.  EXTENSION / RENEWAL")
R += 1

ROW_EXT_YRS     = input_row(R, "Extension / Renewal Term (years)", FMT_INT, "0 = no renewal modeled"); R += 1
ROW_EXT_HAIRCUT = input_row(R, "Renewal Revenue Haircut (%)", FMT_PCT, "Renewal rent as % of base rent  (e.g. 90%)"); R += 1
ROW_PROB_EXT    = input_row(R, "Probability of Renewal (%)", FMT_PCT); R += 1
ROW_PV_RENEWAL  = calc_row(R,  "PV of Renewal (risked, $mm)", FMT_MM, "= Renewal NOI × annuity × discount × prob"); R += 1

R += 1  # blank

# ── SECTION G: RISK ADJUSTMENT ────────────────────────────────────────────────
section_header(R, "G.  RISK ADJUSTMENT  (pipeline / unsigned deals)")
R += 1

ROW_PROB_SIGN   = input_row(R, "Probability of Signing (%)", FMT_PCT, "1.0 = already signed"); R += 1
ROW_PROB_COD    = input_row(R, "Probability of Reaching COD (%)", FMT_PCT, "Execution / construction risk"); R += 1
ROW_RISK_FACTOR = calc_row(R,  "Combined Risk Factor (%)", FMT_PCT, "= Prob Sign × Prob COD"); R += 1

R += 1  # blank

# ── SECTION H: NPV OUTPUT ─────────────────────────────────────────────────────
section_header(R, "H.  NPV OUTPUT  ◀  KEY RESULTS")
R += 1

ROW_PV_BASE     = output_row(R, "PV of Base Lease / Contract ($mm)", FMT_MM, "= Annual NOI × Annuity Factor × Discount Factor"); R += 1
ROW_GROSS_VALUE = output_row(R, "Gross Project Value ($mm)", FMT_MM, "= PV Base + PV Renewal", bold=True, bg=C_MIDBLUE, height=18); R += 1
ROW_UNLEV_NPV   = output_row(R, "Unlevered Project NPV ($mm)", FMT_MM, "= Gross Value - Capex  |  Is this a good project?", bold=True, bg=C_LIGHTBLUE, height=18); R += 1
ROW_LEV_NPV     = output_row(R, "Levered Equity NPV ($mm)", FMT_MM, "= Gross Value - Debt - Equity  |  Equity return", bold=True, bg=C_LIGHTBLUE, height=18); R += 1
ROW_RISK_UNLEV  = output_row(R, "Risked Unlevered NPV ($mm)", FMT_MM, "= Unlevered NPV × Risk Factor", bold=True, bg=C_GREEN, height=20); R += 1
ROW_RISK_LEV    = output_row(R, "Risked Levered NPV ($mm)", FMT_MM, "= Levered NPV × Risk Factor", bold=True, bg=C_GREEN, height=20); R += 1

R += 2  # blank

# ── SECTION I: PORTFOLIO SUMMARY ─────────────────────────────────────────────
ws.row_dimensions[R].height = 20
merge_w(R, 2, 11, "  PORTFOLIO SUMMARY", bg=C_NAVY, fg=C_WHITE, bold=True, size=11, h_align="left")
fill_row(R, 1, 11, C_NAVY)
R += 1

ws.row_dimensions[R].height = 16
merge_w(R, 2, 3, "Metric", bg=C_BLUE, fg=C_WHITE, bold=True, size=9, h_align="left")
for col in PROJECT_COLS:
    c = ws.cell(row=R, column=col)
    c.fill = fill(C_BLUE)
    c.font = font(bold=True, color=C_WHITE, size=9)
    c.alignment = align(h="center")
w(R, 10, "Sum / Note", bg=C_BLUE, fg=C_WHITE, bold=True, size=9)
R += 1

def summary_row(row, label, row_ref, fmt=FMT_MM, bg=C_GREY, bold=False):
    ws.row_dimensions[row].height = 15
    w(row, 2, label, bg=bg, size=9, bold=bold)
    for col in PROJECT_COLS:
        src = ws.cell(row=row_ref, column=col)
        c   = ws.cell(row=row, column=col)
        c.fill  = fill(bg)
        c.font  = font(size=9, bold=bold)
        c.alignment = align(h="center")
        if fmt:
            c.number_format = fmt
        # Reference the output row
        col_letter = get_column_letter(col)
        c.value = f"={col_letter}{row_ref}"

    # SUM formula across project cols
    first = get_column_letter(PROJECT_COLS[0])
    last  = get_column_letter(PROJECT_COLS[-1])
    total_cell = ws.cell(row=row, column=10)
    total_cell.value = f"=SUM({first}{row}:{last}{row})"
    total_cell.fill  = fill(C_ORANGE if bold else bg)
    total_cell.font  = font(bold=bold, size=9)
    total_cell.alignment = align(h="center")
    if fmt:
        total_cell.number_format = fmt

summary_row(R, "Gross Project Value ($mm)",   ROW_GROSS_VALUE, bg=C_LIGHTBLUE); R += 1
summary_row(R, "Unlevered Project NPV ($mm)", ROW_UNLEV_NPV, bg=C_LIGHTBLUE); R += 1
summary_row(R, "Levered Equity NPV ($mm)",    ROW_LEV_NPV, bg=C_LIGHTBLUE); R += 1
summary_row(R, "Risked Unlevered NPV ($mm)",  ROW_RISK_UNLEV, bg=C_GREEN, bold=True); R += 1
summary_row(R, "Risked Levered NPV ($mm)",    ROW_RISK_LEV, bg=C_GREEN, bold=True); R += 1

R += 2

# ─────────────────────────────────────────────────────────────────────────────
# FORMULAS — wire up green calc cells
# ─────────────────────────────────────────────────────────────────────────────

for col in PROJECT_COLS:
    cl = get_column_letter(col)

    def cell(r): return f"{cl}{r}"

    # Annual Revenue: MW × Rev/W  OR  ARR (whichever is non-zero)
    ws.cell(ROW_ANNUAL_REV, col).value = (
        f"=IF({cl}{ROW_MW}*{cl}{ROW_REV_PW}<>0,"
        f"{cl}{ROW_MW}*{cl}{ROW_REV_PW},"
        f"IF({cl}{ROW_ARR}<>\"\",{cl}{ROW_ARR},0))"
    )

    # Realized Revenue = Annual × Utilization
    ws.cell(ROW_REALIZED, col).value = (
        f"=IF({cl}{ROW_UTIL}<>\"\",{cl}{ROW_ANNUAL_REV}*{cl}{ROW_UTIL},{cl}{ROW_ANNUAL_REV})"
    )

    # Annual NOI
    ws.cell(ROW_ANNUAL_NOI, col).value = (
        f"={cl}{ROW_REALIZED}*{cl}{ROW_NOI_MARGIN}"
    )

    # Cloud EBITDA
    ws.cell(ROW_CLOUD_EBITDA, col).value = (
        f"={cl}{ROW_REALIZED}-{cl}{ROW_POWER}-{cl}{ROW_DC_OPEX}-{cl}{ROW_SW_SUPPORT}"
    )

    # Cloud FCF
    ws.cell(ROW_CLOUD_FCF, col).value = (
        f"={cl}{ROW_CLOUD_EBITDA}-{cl}{ROW_INTEREST}"
    )

    # Years to COD
    ws.cell(ROW_YRS_TO_COD, col).value = f"={cl}{ROW_COD}-2026"

    # Annuity factor: (1-(1+r)^-n)/r
    ws.cell(ROW_ANN_FACTOR, col).value = (
        f"=IF({cl}{ROW_DISC}=0,{cl}{ROW_TERM},"
        f"(1-(1+{cl}{ROW_DISC})^(-{cl}{ROW_TERM}))/{cl}{ROW_DISC})"
    )

    # Discount factor: 1/(1+r)^t
    ws.cell(ROW_DISC_FACTOR, col).value = (
        f"=1/(1+{cl}{ROW_DISC})^{cl}{ROW_YRS_TO_COD}"
    )

    # Equity invested
    ws.cell(ROW_EQUITY, col).value = (
        f"={cl}{ROW_CAPEX}*(1-{cl}{ROW_LTC})"
    )

    # GPU replacement
    ws.cell(ROW_REPL_COST, col).value = (
        f"={cl}{ROW_CAPEX}*{cl}{ROW_REPL_PCT}"
    )
    ws.cell(ROW_REPL_YR, col).value = (
        f"=IF({cl}{ROW_GPU_LIFE}=\"\",\"\",{cl}{ROW_COD}+{cl}{ROW_GPU_LIFE})"
    )
    ws.cell(ROW_PV_REPL, col).value = (
        f"=IF({cl}{ROW_GPU_LIFE}=\"\",0,"
        f"{cl}{ROW_REPL_COST}/(1+{cl}{ROW_DISC})^({cl}{ROW_REPL_YR}-2026))"
    )

    # PV of renewal
    ws.cell(ROW_PV_RENEWAL, col).value = (
        f"=IF({cl}{ROW_EXT_YRS}=0,0,"
        f"{cl}{ROW_ANNUAL_NOI}*{cl}{ROW_EXT_HAIRCUT}"
        f"*IF({cl}{ROW_DISC}=0,{cl}{ROW_EXT_YRS},(1-(1+{cl}{ROW_DISC})^(-{cl}{ROW_EXT_YRS}))/{cl}{ROW_DISC})"
        f"*{cl}{ROW_DISC_FACTOR}"
        f"*(1+{cl}{ROW_DISC})^(-{cl}{ROW_TERM})"
        f"*{cl}{ROW_PROB_EXT})"
    )

    # Risk factor
    ws.cell(ROW_RISK_FACTOR, col).value = (
        f"={cl}{ROW_PROB_SIGN}*{cl}{ROW_PROB_COD}"
    )

    # PV of base lease/contract
    # For colo: NOI × annuity × discount
    # For cloud: FCF × annuity × discount - PV replacement
    ws.cell(ROW_PV_BASE, col).value = (
        f'=IF(UPPER({cl}{ROW_TYPE})="CLOUD",'
        f'{cl}{ROW_CLOUD_FCF}*{cl}{ROW_ANN_FACTOR}*{cl}{ROW_DISC_FACTOR}-{cl}{ROW_PV_REPL},'
        f'{cl}{ROW_ANNUAL_NOI}*{cl}{ROW_ANN_FACTOR}*{cl}{ROW_DISC_FACTOR})'
    )

    # Gross project value
    ws.cell(ROW_GROSS_VALUE, col).value = (
        f"={cl}{ROW_PV_BASE}+{cl}{ROW_PV_RENEWAL}"
    )

    # Unlevered NPV
    ws.cell(ROW_UNLEV_NPV, col).value = (
        f"={cl}{ROW_GROSS_VALUE}-{cl}{ROW_CAPEX}"
    )

    # Levered NPV
    ws.cell(ROW_LEV_NPV, col).value = (
        f"={cl}{ROW_GROSS_VALUE}-{cl}{ROW_DEBT}-{cl}{ROW_EQUITY}"
    )

    # Risked NPVs
    ws.cell(ROW_RISK_UNLEV, col).value = (
        f"={cl}{ROW_UNLEV_NPV}*{cl}{ROW_RISK_FACTOR}"
    )
    ws.cell(ROW_RISK_LEV, col).value = (
        f"={cl}{ROW_LEV_NPV}*{cl}{ROW_RISK_FACTOR}"
    )

# ─────────────────────────────────────────────────────────────────────────────
# PRE-FILL: Example projects from your research
# ─────────────────────────────────────────────────────────────────────────────

PROJECTS_DATA = [
    # Col D: CIFR Barber Lake — signed 10-yr NNN, Fluidstack/Google, $3.83B total
    {
        "type": "Colo", "company": "CIFR", "name": "Barber Lake (Fluidstack)",
        "status": "Signed 10-yr NNN; Fluidstack/Google-backed; Phase I Sep 2026 → full ramp Jan 2027",
        "mw": 207, "rev_pw": None, "arr": 383, "util": 1.0,   # $3.83B / 10yr
        "noi_margin": 0.85,
        "power": None, "dc_opex": None, "sw": None, "interest": None,
        "cod": 2027, "term": 10, "disc": 0.10,
        "capex": 400, "ltc": 0.85, "debt": 340,
        "gpu_life": None, "repl_pct": None,
        "ext_yrs": 5, "ext_haircut": 0.90, "prob_ext": 0.65,
        "prob_sign": 1.0, "prob_cod": 0.90,
    },
    # Col E: CIFR Black Pearl — signed 15-yr NNN, Amazon Data Services, $5.5B total
    {
        "type": "Colo", "company": "CIFR", "name": "Black Pearl (AWS)",
        "status": "Signed 15-yr NNN; Amazon Data Services; Phase I Q4 2026 → full ramp Q1 2027",
        "mw": 216, "rev_pw": None, "arr": 367, "util": 1.0,   # $5.5B / 15yr
        "noi_margin": 0.85,
        "power": None, "dc_opex": None, "sw": None, "interest": None,
        "cod": 2027, "term": 15, "disc": 0.10,
        "capex": 430, "ltc": 0.85, "debt": 365,
        "gpu_life": None, "repl_pct": None,
        "ext_yrs": 5, "ext_haircut": 0.90, "prob_ext": 0.65,
        "prob_sign": 1.0, "prob_cod": 0.90,
    },
    # Col F: CIFR Stingray — signed long-term NNN, Amazon Data Services, $2.06B total
    {
        "type": "Colo", "company": "CIFR", "name": "Stingray (AWS)",
        "status": "Signed long-term NNN; Amazon Data Services; earthwork+substation Q1 2026; H1 2026 COD",
        "mw": 70, "rev_pw": None, "arr": 137, "util": 1.0,    # ~$2.06B / 15yr
        "noi_margin": 0.85,
        "power": None, "dc_opex": None, "sw": None, "interest": None,
        "cod": 2026, "term": 15, "disc": 0.10,
        "capex": 953, "ltc": 0.85, "debt": 810,               # $810M project financing = debt
        "gpu_life": None, "repl_pct": None,
        "ext_yrs": 5, "ext_haircut": 0.90, "prob_ext": 0.65,
        "prob_sign": 1.0, "prob_cod": 0.92,
    },
    # Col G: CIFR Colchis — pipeline, 1 GW site, West Texas, AEP dual interconnect
    {
        "type": "Colo", "company": "CIFR", "name": "Colchis (W. Texas)",
        "status": "Pipeline; 1 GW site; majority JV stake; AEP dual interconnect executed; no lease signed",
        "mw": 500, "rev_pw": 1.50, "arr": None, "util": 1.0,  # Phase 1 ~500 MW estimate
        "noi_margin": 0.85,
        "power": None, "dc_opex": None, "sw": None, "interest": None,
        "cod": 2029, "term": 15, "disc": 0.10,
        "capex": 750, "ltc": 0.80, "debt": 600,
        "gpu_life": None, "repl_pct": None,
        "ext_yrs": 5, "ext_haircut": 0.85, "prob_ext": 0.55,
        "prob_sign": 0.35, "prob_cod": 0.75,
    },
    # Col H: CIFR Ulysses — pipeline, Ohio, 200 MW, AEP Ohio / future PJM
    {
        "type": "Colo", "company": "CIFR", "name": "Ulysses (Ohio)",
        "status": "Pipeline; site acquired Dec 2025; AEP Ohio / future PJM; Q4 2027 est.",
        "mw": 200, "rev_pw": 1.50, "arr": None, "util": 1.0,
        "noi_margin": 0.85,
        "power": None, "dc_opex": None, "sw": None, "interest": None,
        "cod": 2028, "term": 15, "disc": 0.10,
        "capex": 300, "ltc": 0.80, "debt": 240,
        "gpu_life": None, "repl_pct": None,
        "ext_yrs": 5, "ext_haircut": 0.85, "prob_ext": 0.55,
        "prob_sign": 0.40, "prob_cod": 0.78,
    },
    # Col I: CIFR McLennan — pipeline, Central Texas, 500 MW, ERCOT queue
    {
        "type": "Colo", "company": "CIFR", "name": "McLennan (Central TX)",
        "status": "Pipeline; option exercised Feb 2026; ERCOT interconnect queue; est. 2028",
        "mw": 500, "rev_pw": 1.50, "arr": None, "util": 1.0,
        "noi_margin": 0.85,
        "power": None, "dc_opex": None, "sw": None, "interest": None,
        "cod": 2028, "term": 15, "disc": 0.10,
        "capex": 750, "ltc": 0.80, "debt": 600,
        "gpu_life": None, "repl_pct": None,
        "ext_yrs": 5, "ext_haircut": 0.85, "prob_ext": 0.55,
        "prob_sign": 0.35, "prob_cod": 0.75,
    },
]

def set_val(row, col, val, fmt=None):
    c = ws.cell(row=row, column=col)
    if val is not None:
        c.value = val
    if fmt:
        c.number_format = fmt

for i, p in enumerate(PROJECTS_DATA):
    col = PROJECT_COLS[i]

    set_val(ROW_TYPE,        col, p["type"])
    set_val(ROW_COMPANY,     col, p["company"])
    set_val(ROW_PROJECT_NAME,col, p["name"])
    set_val(ROW_STATUS,      col, p["status"])

    if p["mw"]:     set_val(ROW_MW,     col, p["mw"],     FMT_INT)
    if p["rev_pw"]: set_val(ROW_REV_PW, col, p["rev_pw"], '"$"#,##0.000')
    if p["arr"]:    set_val(ROW_ARR,    col, p["arr"],     FMT_MM)
    set_val(ROW_UTIL,       col, p["util"],       FMT_PCT)
    set_val(ROW_NOI_MARGIN, col, p["noi_margin"], FMT_PCT)

    if p["power"]:    set_val(ROW_POWER,      col, p["power"],    FMT_MM)
    if p["dc_opex"]:  set_val(ROW_DC_OPEX,    col, p["dc_opex"],  FMT_MM)
    if p["sw"]:       set_val(ROW_SW_SUPPORT,  col, p["sw"],       FMT_MM)
    if p["interest"]: set_val(ROW_INTEREST,    col, p["interest"], FMT_MM)

    set_val(ROW_COD,  col, p["cod"],  FMT_INT)
    set_val(ROW_TERM, col, p["term"], FMT_INT)
    set_val(ROW_DISC, col, p["disc"], FMT_PCT)

    set_val(ROW_CAPEX, col, p["capex"], FMT_MM)
    set_val(ROW_LTC,   col, p["ltc"],   FMT_PCT)
    set_val(ROW_DEBT,  col, p["debt"],  FMT_MM)

    if p["gpu_life"]:  set_val(ROW_GPU_LIFE,  col, p["gpu_life"],  FMT_INT)
    if p["repl_pct"]:  set_val(ROW_REPL_PCT,  col, p["repl_pct"],  FMT_PCT)

    set_val(ROW_EXT_YRS,    col, p["ext_yrs"],    FMT_INT)
    set_val(ROW_EXT_HAIRCUT,col, p["ext_haircut"], FMT_PCT)
    set_val(ROW_PROB_EXT,   col, p["prob_ext"],    FMT_PCT)

    set_val(ROW_PROB_SIGN, col, p["prob_sign"], FMT_PCT)
    set_val(ROW_PROB_COD,  col, p["prob_cod"],  FMT_PCT)

# ─────────────────────────────────────────────────────────────────────────────
# FREEZE PANES & SHEET SETTINGS
# ─────────────────────────────────────────────────────────────────────────────

ws.freeze_panes = "D8"   # freeze row headers and first few rows
ws.sheet_view.showGridLines = False

# ─────────────────────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────────────────────

output_path = "/home/user/claude-code/npv_model.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
